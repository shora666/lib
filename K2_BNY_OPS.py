#! Python3

import os
import pandas as pd
import datetime
import openpyxl # version 2.5.14
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border, NamedStyle, numbers
from datetime import timedelta

import shutil
import sys

def BNY_FX(mapping_loc, crystal_dest_loc):
    #procedure to update prior FX 
    read_mapping = openpyxl.load_workbook(mapping_loc)
    mapping_sheet2 = read_mapping['Sheet2']
    #update current FX with values from portfolio crystalization file
    # start with portfolio positions
    crystalization_file = openpyxl.load_workbook(crystal_dest_loc)
    portfolio_position = crystalization_file['Portfolio Positions']
    for cry_row in range(2, portfolio_position.max_row + 1):
        for row in range(2, mapping_sheet2.max_row + 1):
            if portfolio_position['B{}'.format(cry_row)].value == mapping_sheet2['E{}'.format(row)].value and portfolio_position['C{}'.format(cry_row)].value == mapping_sheet2['D{}'.format(row)].value:
                mapping_sheet2['R{}'.format(row)].value = portfolio_position['G{}'.format(cry_row)].value
    # continue with portfolio accounts
    portfolio_accounts = crystalization_file['Portfolio Accounts']
    for cry_row in range(2, portfolio_accounts.max_row + 1):
        for row in range(2, mapping_sheet2.max_row + 1):
            if portfolio_accounts['B{}'.format(cry_row)].value == mapping_sheet2['D{}'.format(row)].value:
                mapping_sheet2['R{}'.format(row)].value = portfolio_accounts['V{}'.format(cry_row)].value
    read_mapping.save(mapping_loc)
    crystalization_file.save(crystal_dest_loc)


def BNY_Close_Value_Update(mapping_loc):
    #procedure to update closing value in BNY mapping file
    read_mapping = openpyxl.load_workbook(mapping_loc)
    mapping_sheet2 = read_mapping['Sheet2']
    for row in range(2, mapping_sheet2.max_row + 1):
        mapping_sheet2['N{}'.format(row)].value = mapping_sheet2['I{}'.format(row)].value * mapping_sheet2['J{}'.format(row)].value
    read_mapping.save(mapping_loc)


def K2_BNY_OPS(working_date,folder_path,input_file,os):
    # convert txt to xlsx file and save
    #text_data_df_file = input('Please enter input file.\n')
    text_data_df_file = folder_path +"/"+input_file
    text_data_df = pd.read_csv(text_data_df_file, sep = '|')
    text_data_df_maxrow = len(text_data_df)
    today = datetime.date.today()
    todayStr = today.strftime('%Y-%m-%d')
    pd_file_date = text_data_df.columns[1]
    pd_file_datetime = datetime.datetime.strptime(pd_file_date, '%m/%d/%Y')
    pd_file_datetime_formatted = pd_file_datetime.strftime('%Y-%m-%d')
    if (os=='Windows'):
        text_to_data = r'G:/Shared drives/Operations/K2/BNY - Daily/' + pd_file_datetime_formatted + ' text to data ' + todayStr + '.xlsx'
    else:
        text_to_data = f'{folder_path}/'  + pd_file_datetime_formatted + ' text to data ' + todayStr + '.xlsx'
    text_data_df.to_excel(text_to_data ,'Sheet1') #can just use df to read -> keep in case

    # openpyxl to forward fill merged cells
    read_text_to_data = openpyxl.load_workbook(text_to_data)
    text_to_data_sheet = read_text_to_data['Sheet1']

    merged_ranges = text_to_data_sheet.merged_cells.ranges
    counter = len(merged_ranges)
    while counter != 0:
        text_to_data_sheet.unmerge_cells(str(merged_ranges[0]))
        counter -= 1

    for rownum in range(3, text_data_df_maxrow):
        if text_to_data_sheet['A{}'.format(rownum)].value == None:
            text_to_data_sheet['A{}'.format(rownum)].value = text_to_data_sheet['A{}'.format(rownum-1)].value

    # change to number format
    for rownum in range(3, text_data_df_maxrow + 1):
        text_to_data_sheet['A{}'.format(rownum)].value = int(text_to_data_sheet['A{}'.format(rownum)].value)
        text_to_data_sheet['C{}'.format(rownum)].value = float(text_to_data_sheet['C{}'.format(rownum)].value)
        text_to_data_sheet['D{}'.format(rownum)].value = float(text_to_data_sheet['D{}'.format(rownum)].value)
        text_to_data_sheet['E{}'.format(rownum)].value = float(text_to_data_sheet['E{}'.format(rownum)].value)
        text_to_data_sheet['F{}'.format(rownum)].value = float(text_to_data_sheet['F{}'.format(rownum)].value)
        text_to_data_sheet['I{}'.format(rownum)].value = float(text_to_data_sheet['I{}'.format(rownum)].value)
        text_to_data_sheet['J{}'.format(rownum)].value = float(text_to_data_sheet['J{}'.format(rownum)].value)
        text_to_data_sheet['K{}'.format(rownum)].value = float(text_to_data_sheet['K{}'.format(rownum)].value)
        text_to_data_sheet['L{}'.format(rownum)].value = float(text_to_data_sheet['L{}'.format(rownum)].value)

    # edit txt file for easy transfer

    ###

    file_date_str = text_to_data_sheet['K1'].value
    file_date_datetime = datetime.datetime.strptime(file_date_str, '%m/%d/%Y')
    file_date_formatted = file_date_datetime.date()
    file_date_ext = file_date_formatted.strftime('%Y-%m-%d')
    #now = datetime.now()
    #nowstr = now.strftime("%d %m %Y %H %M %S")

    # rename text to data file
    if (os=='Windows'):
        new_text_to_data = r'G:/Shared drives/Operations/K2/BNY - Daily/' + file_date_ext + ' text to data ' + todayStr + '.xlsx'
    else:
        new_text_to_data =  f'{folder_path}/' + file_date_ext + ' text to data ' + todayStr + '.xlsx'
    #os.rename(text_to_data, new_text_to_data)

    ###


    # open mapping file - to be edited later
    if (os=='Windows'):
        mapping = r'G:/Shared drives/Operations/K2/BNY - Daily/BNY Mapping.xlsx'
        mapping_backup = r'G:/Shared drives/Operations/K2/BNY - Daily/' + 'BNY Mapping Backup pre ' + file_date_ext + '.xlsx'
    else:
        mapping = f'{folder_path}/mapping/BNY Mapping.xlsx'
        mapping_backup = f'{folder_path}/backup/' + 'BNY Mapping Backup pre ' + file_date_ext + '.xlsx'

    read_mapping = openpyxl.load_workbook(mapping)
    mapping_sheet2 = read_mapping['Sheet2']
    mapping_sheet3 = read_mapping['Sheet3']


    print("The last crystallization date ran was " + mapping_sheet3['B2'].value.strftime('%m-%d-%y') + ".")
    continue_run = True
    cry_start_date = mapping_sheet3['B2'].value
    cry_start_date += timedelta(days=1)

    #run_decision = input("Would you like to continue? (y/n)\n")
    run_decision = 'y'
    while continue_run:
        if run_decision == 'y':
            continue_run = False
        elif run_decision == 'n':
            sys.exit()

    shutil.copyfile(mapping, mapping_backup)
    for row in range(2, mapping_sheet2.max_row + 1):   #copy current fields in mapping to prior fields
        mapping_sheet2['S{}'.format(row)].value = mapping_sheet2['R{}'.format(row)].value # FX
        mapping_sheet2['T{}'.format(row)].value = mapping_sheet2['G{}'.format(row)].value # Cry Date
        mapping_sheet2['O{}'.format(row)].value = mapping_sheet2['I{}'.format(row)].value # NAV per share
        mapping_sheet2['P{}'.format(row)].value = mapping_sheet2['J{}'.format(row)].value # Shares
        mapping_sheet2['Q{}'.format(row)].value = mapping_sheet2['N{}'.format(row)].value # Value
 

    # update share value on investment pools prior to running
    ##days = [1, 2, 3]
    ##if file_date_datetime.day in days and (file_date_datetime.weekday() < 6):
    ##    print("Update needed on shares as it is the first of the month or after.\n")
    ##    alternative_shares = float(input("Please enter the shares for the BNY Alternative Strategies Portfolio.\n"))
    ##    mapping_sheet2['J48'].value = alternative_shares
    ##    ls_shares = float(input("Please enter the shares for the BNY LS Portfolio.\n"))
    ##    mapping_sheet2['J49'].value = ls_shares

    #######################################################################################################################
    # create exception report - acct #s not in text file
    account_number_list = []

    if (os=='Windows'):
        missing_account_directory = r'G:/Shared drives/Operations/K2/BNY - Daily/'
    else:
        missing_account_directory = f'{folder_path}/processing'
    missing_account = Workbook()
    missing_account_sheet = missing_account.create_sheet('Missing Accounts', 0)
    missing_account_sheet['A1'].value = 'InvestmentID'
    missing_account_sheet['B2'].value = 'Investment Name'

    for acctnum in range(2, mapping_sheet2.max_row):
        account_number_list.append(mapping_sheet2['A{}'.format(acctnum)].value)

    filler = 2
    missing_counter = 0
    for row in range(3, text_data_df_maxrow):
        if text_to_data_sheet['A{}'.format(row)].value not in account_number_list:
            print(str(text_to_data_sheet['A{}'.format(row)].value) + ' not in mapping file.')
            missing_account_sheet['A{}'.format(filler)].value = text_to_data_sheet['A{}'.format(row)].value
            missing_account_sheet['B{}'.format(filler)].value = text_to_data_sheet['B{}'.format(row)].value
            missing_account_sheet['C{}'.format(filler)].value = text_to_data_sheet['C{}'.format(row)].value
            missing_account_sheet['D{}'.format(filler)].value = text_to_data_sheet['D{}'.format(row)].value
            missing_account_sheet['E{}'.format(filler)].value = text_to_data_sheet['E{}'.format(row)].value
            missing_account_sheet['F{}'.format(filler)].value = text_to_data_sheet['F{}'.format(row)].value
            missing_account_sheet['G{}'.format(filler)].value = text_to_data_sheet['G{}'.format(row)].value
            missing_account_sheet['H{}'.format(filler)].value = text_to_data_sheet['H{}'.format(row)].value
            missing_account_sheet['I{}'.format(filler)].value = text_to_data_sheet['I{}'.format(row)].value
            missing_account_sheet['J{}'.format(filler)].value = text_to_data_sheet['J{}'.format(row)].value
            missing_account_sheet['K{}'.format(filler)].value = text_to_data_sheet['K{}'.format(row)].value
            missing_account_sheet['L{}'.format(filler)].value = text_to_data_sheet['L{}'.format(row)].value
            filler += 1
            missing_counter += 1

    if missing_counter == 0:
        print('All account numbers were found in mapping file.\n')
    else:
        missing_account.save(missing_account_directory + 'K2 BNY Missing Accounts ' + todayStr + '.xlsx')
        sys.exit() # stop for exception report

    # exception report and bugs fixed
    # bugs
    # 1. merged cells
    # 2. fill cells after unmerged 
    #--------------------------------------------------------------------------------------------------

    # txt to csv edit
    text_to_data_sheet.insert_cols(3)
    text_to_data_sheet.insert_cols(3)

    text_to_data_sheet['C2'].value = 'Investor Account'
    text_to_data_sheet['D2'].value = 'Investment Account'
    text_to_data_sheet['O2'].value = 'Portfolio Column'

    for row in range(3, text_data_df_maxrow + 1):
        if text_to_data_sheet['B{}'.format(row)].value == '0':
            for account_val in range(2, mapping_sheet2.max_row):
                if mapping_sheet2['A{}'.format(account_val)].value == text_to_data_sheet['A{}'.format(row)].value:
                    text_to_data_sheet['C{}'.format(row)].value = mapping_sheet2['E{}'.format(account_val)].value
                    text_to_data_sheet['D{}'.format(row)].value = mapping_sheet2['D{}'.format(account_val)].value
                    text_to_data_sheet['O{}'.format(row)].value = mapping_sheet2['F{}'.format(account_val)].value
        elif text_to_data_sheet['B{}'.format(row)].value != '0':
            for account_val in range(2, mapping_sheet2.max_row):
                if mapping_sheet2['A{}'.format(account_val)].value == str(text_to_data_sheet['A{}'.format(row)].value) + text_to_data_sheet['B{}'.format(row)].value:
                    text_to_data_sheet['C{}'.format(row)].value = mapping_sheet2['E{}'.format(account_val)].value
                    text_to_data_sheet['D{}'.format(row)].value = mapping_sheet2['D{}'.format(account_val)].value
                    text_to_data_sheet['O{}'.format(row)].value = mapping_sheet2['F{}'.format(account_val)].value

    read_text_to_data.save(filename = text_to_data)

    ####################################################################
    # Valuation File Start

    if (os=='Windows'):
        valuation_file = r'G:/Shared drives/Operations/K2/BNY - Daily/Valuation Template.xlsx'
    else:
        valuation_file = f'{folder_path}/templates/Valuation Template.xlsx'

    valuation = openpyxl.load_workbook(valuation_file)
    mgr_accts_perf = valuation['Mgr Accts Performance']

    mgr_accts_perf.delete_rows(2, 1000) # clean out file

    valuation_counter = 2
    for row in range(3, text_data_df_maxrow):
        if text_to_data_sheet['G{}'.format(row)].value != 0 and text_to_data_sheet['B{}'.format(row)].value == '0' and text_to_data_sheet['C{}'.format(row)].value != None:
            # vlookup for name
            for row_map in range(2, mapping_sheet2.max_row):
                if text_to_data_sheet['A{}'.format(row)].value == mapping_sheet2['A{}'.format(row_map)].value:
                    mgr_accts_perf['A{}'.format(valuation_counter)].value = mapping_sheet2['D{}'.format(row_map)].value
                else:
                    continue
            mgr_accts_perf['B{}'.format(valuation_counter)].value = file_date_formatted # date
            mgr_accts_perf['B{}'.format(valuation_counter)].number_format = numbers.FORMAT_DATE_XLSX14
            mgr_accts_perf['D{}'.format(valuation_counter)].value = text_to_data_sheet['G{}'.format(row)].value # nav
            mgr_accts_perf['D{}'.format(valuation_counter)].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            mgr_accts_perf['E{}'.format(valuation_counter)].value = 'TRUE'  # final
            valuation_counter += 1

    # remove blank name
    for row in range(2, valuation_counter):
        if not mgr_accts_perf['A{}'.format(row)].value:
            mgr_accts_perf.delete_rows(row, 1)

    mgr_accts_perf.delete_rows(mgr_accts_perf.max_row + 1, 50)

    #--------------------------------------------------------------------------------------------------

    #######################################################################
    # Crystallization File Start

    if (os=='Windows'):
        crystal_file = r'G:/Shared drives/Operations/K2/BNY - Daily/Crystallization Template.xlsx'
    else:
        crystal_file = f'{folder_path}/templates/Crystallization Template.xlsx'

    crystal = openpyxl.load_workbook(crystal_file)

    # open all worksheets in crystal file
    portfolio_crystal = crystal['Portfolio Crystallization']
    portfolio_positions = crystal['Portfolio Positions']
    portfolio_accounts = crystal['Portfolio Accounts']

    # start by clearing out file for processing
    portfolio_crystal.delete_rows(2, 1000)
    portfolio_positions.delete_rows(2, 1000)
    portfolio_accounts.delete_rows(2, 1000)

    # start working on portfolio crystallization sheet
    idx = (pd_file_datetime.weekday() + 1) % 7
    sat = pd_file_datetime - datetime.timedelta(7+idx-6)

    portfolio_crystal_funds = []

    for row in range(2, mapping_sheet2.max_row):
        if mapping_sheet2['F{}'.format(row)].value and mapping_sheet2['F{}'.format(row)].value not in portfolio_crystal_funds:
            portfolio_crystal_funds.append(mapping_sheet2['F{}'.format(row)].value)

    portfolio_crystal_counter = 2
    for index, value in enumerate(portfolio_crystal_funds):
        portfolio_crystal['A{}'.format(portfolio_crystal_counter)].value = value + ' - ' + file_date_ext # crystallization reference
        portfolio_crystal['C{}'.format(portfolio_crystal_counter)].value = value # Portfolio Long Name
        #if pd_file_datetime.weekday() == 0:
        #    portfolio_crystal['D{}'.format(portfolio_crystal_counter)].value = sat # Starting Date
        #    portfolio_crystal['D{}'.format(portfolio_crystal_counter)].number_format = numbers.FORMAT_DATE_XLSX14
        #else:
        #    portfolio_crystal['D{}'.format(portfolio_crystal_counter)].value = file_date_formatted # starting date
        #    portfolio_crystal['D{}'.format(portfolio_crystal_counter)].number_format = numbers.FORMAT_DATE_XLSX14
        portfolio_crystal['D{}'.format(portfolio_crystal_counter)].value = cry_start_date # Starting Date
        portfolio_crystal['D{}'.format(portfolio_crystal_counter)].number_format = numbers.FORMAT_DATE_XLSX14
        portfolio_crystal['E{}'.format(portfolio_crystal_counter)].value = file_date_formatted # Ending Date
        portfolio_crystal['E{}'.format(portfolio_crystal_counter)].number_format = numbers.FORMAT_DATE_XLSX14
        portfolio_crystal['F{}'.format(portfolio_crystal_counter)].value = file_date_formatted # Statement Date
        portfolio_crystal['F{}'.format(portfolio_crystal_counter)].number_format = numbers.FORMAT_DATE_XLSX14
        portfolio_crystal['G{}'.format(portfolio_crystal_counter)].value = 'Final' # Status
        portfolio_crystal_counter += 1

    # start working on portfolio positions and port accounts 
    portfolio_positions_counter = 2
    portfolio_accounts_counter = 2

    ##adj_opening_multi = 0
    ##adj_opening_ls = 0

    for row in range(3, text_data_df_maxrow):
        if text_to_data_sheet['B{}'.format(row)].value == '0' and text_to_data_sheet['D{}'.format(row)].value:
            if text_to_data_sheet['D{}'.format(row)].value != None:
                if 'Managed' not in text_to_data_sheet['D{}'.format(row)].value and 'K2 HOLDINGS INVESTMENT CORPORATION' not in text_to_data_sheet['D{}'.format(row)].value:
                    if text_to_data_sheet['H{}'.format(row)].value > 0:
                        portfolio_positions['A{}'.format(portfolio_positions_counter)].value = text_to_data_sheet['O{}'.format(row)].value + ' - ' + file_date_ext # crystal reference
                        portfolio_positions['B{}'.format(portfolio_positions_counter)].value = text_to_data_sheet['C{}'.format(row)].value # investor acct
                        portfolio_positions['C{}'.format(portfolio_positions_counter)].value = text_to_data_sheet['D{}'.format(row)].value # ivstment acct
                        portfolio_positions['D{}'.format(portfolio_positions_counter)].value = text_to_data_sheet['H{}'.format(row)].value # units closing
                        portfolio_positions['D{}'.format(portfolio_positions_counter)].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                        portfolio_positions['E{}'.format(portfolio_positions_counter)].value = text_to_data_sheet['G{}'.format(row)].value # closing nav
                        portfolio_positions['E{}'.format(portfolio_positions_counter)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                        portfolio_positions['F{}'.format(portfolio_positions_counter)].value = text_to_data_sheet['E{}'.format(row)].value # investment closing bal
                        portfolio_positions['F{}'.format(portfolio_positions_counter)].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                        portfolio_positions['G{}'.format(portfolio_positions_counter)].value = 1 # invstor closing fx rate
                        portfolio_positions['H{}'.format(portfolio_positions_counter)].value = text_to_data_sheet['E{}'.format(row)].value # invstor closing bal
                        portfolio_positions['H{}'.format(portfolio_positions_counter)].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                        portfolio_positions['I{}'.format(portfolio_positions_counter)].value = 1 # port closing fx
                        portfolio_positions['J{}'.format(portfolio_positions_counter)].value = text_to_data_sheet['E{}'.format(row)].value # port closing bal
                        portfolio_positions['J{}'.format(portfolio_positions_counter)].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                        portfolio_positions_counter += 1
        elif text_to_data_sheet['B{}'.format(row)].value != '0' and text_to_data_sheet['D{}'.format(row)].value:
            for acctnum in range(2, mapping_sheet2.max_row):
                if text_to_data_sheet['D{}'.format(row)].value == mapping_sheet2['D{}'.format(acctnum)].value:
                    portfolio_accounts['A{}'.format(portfolio_accounts_counter)].value = mapping_sheet2['F{}'.format(acctnum)].value + ' - ' + file_date_ext  # crystal ref
                    portfolio_accounts['B{}'.format(portfolio_accounts_counter)].value = text_to_data_sheet['D{}'.format(row)].value # Port Acct
                    portfolio_accounts['C{}'.format(portfolio_accounts_counter)].value = None # opening Units
                    portfolio_accounts['D{}'.format(portfolio_accounts_counter)].value = mapping_sheet2['I{}'.format(acctnum)].value # Opening Nav
                    portfolio_accounts['E{}'.format(portfolio_accounts_counter)].value = None # Opening FX Rate
                    portfolio_accounts['F{}'.format(portfolio_accounts_counter)].value = mapping_sheet2['I{}'.format(acctnum)].value * text_to_data_sheet['H{}'.format(row)].value # Adj Opening Bal
                    portfolio_accounts['G{}'.format(portfolio_accounts_counter)].value = text_to_data_sheet['E{}'.format(row)].value - portfolio_accounts['F{}'.format(portfolio_accounts_counter)].value # Invst PnL
                    portfolio_accounts['G{}'.format(portfolio_accounts_counter)].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                    portfolio_accounts['H{}'.format(portfolio_accounts_counter)].value = 0 # FX PnL
                    portfolio_accounts['I{}'.format(portfolio_accounts_counter)].value = 0 # FX Hedge PnL
                    portfolio_accounts['J{}'.format(portfolio_accounts_counter)].value = 0 # Interest income
                    portfolio_accounts['K{}'.format(portfolio_accounts_counter)].value = 0 # int expense
                    portfolio_accounts['L{}'.format(portfolio_accounts_counter)].value = 0 # admin fe
                    portfolio_accounts['M{}'.format(portfolio_accounts_counter)].value = 0 # custody fee
                    portfolio_accounts['N{}'.format(portfolio_accounts_counter)].value = 0 # misc expense
                    portfolio_accounts['O{}'.format(portfolio_accounts_counter)].value = 0 # acct expenses
                    portfolio_accounts['P{}'.format(portfolio_accounts_counter)].value = 0 # management fee
                    portfolio_accounts['Q{}'.format(portfolio_accounts_counter)].value = 0 # perf fee
                    portfolio_accounts['R{}'.format(portfolio_accounts_counter)].value = 0 # acct fees
                    portfolio_accounts['S{}'.format(portfolio_accounts_counter)].value = text_to_data_sheet['E{}'.format(row)].value # closing bal
                    portfolio_accounts['S{}'.format(portfolio_accounts_counter)].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                    portfolio_accounts['T{}'.format(portfolio_accounts_counter)].value = text_to_data_sheet['H{}'.format(row)].value # closing units
                    portfolio_accounts['T{}'.format(portfolio_accounts_counter)].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                    portfolio_accounts['U{}'.format(portfolio_accounts_counter)].value = text_to_data_sheet['G{}'.format(row)].value # closing nav
                    portfolio_accounts['U{}'.format(portfolio_accounts_counter)].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                    portfolio_accounts['V{}'.format(portfolio_accounts_counter)].value = 1
            
    ##                if 'Long Short' in portfolio_accounts['B{}'.format(portfolio_accounts_counter)].value:
    ##                    adj_opening_ls += portfolio_accounts['G{}'.format(portfolio_accounts_counter)].value
    ##                elif 'Alternative Strategies' in portfolio_accounts['B{}'.format(portfolio_accounts_counter)].value:
    ##                    adj_opening_multi += portfolio_accounts['G{}'.format(portfolio_accounts_counter)].value
                    
                    portfolio_accounts_counter += 1

    # add in sum for port accts - Investment Pool

    # initialize investment pools with their PnL
    investment_pool_funds = {}

    # create a dictionary correlating Investment Pool Name: Investment Pool PnL
    for row in range(2, portfolio_positions.max_row + 1):
        if portfolio_positions['B{}'.format(row)].value and portfolio_positions['B{}'.format(row)].value not in investment_pool_funds:
            investment_pool_funds.setdefault(str(portfolio_positions['B{}'.format(row)].value), 0)

    # comparing the id's for portfolio positions in compiled file sheet with mapping sheet
    # we will see if the Investor Name for XML positions is equal, thus adding PnL to the dictionary key
    for row in range(2, portfolio_accounts.max_row + 1):
        for mapping_row in range(2, mapping_sheet2.max_row):
            if portfolio_accounts['B{}'.format(row)].value == mapping_sheet2['D{}'.format(mapping_row)].value:
                investment_pool_funds[str(mapping_sheet2['E{}'.format(mapping_row)].value)[:-8] + 'Investment Pool'] += portfolio_accounts['G{}'.format(row)].value

    for num, key in enumerate(investment_pool_funds):
        current_row = mapping_sheet2.max_row - len(investment_pool_funds) + 1 + num
        print ("Current Row: ",current_row)
        portfolio_accounts['A{}'.format(portfolio_accounts_counter)].value = str(mapping_sheet2['F{}'.format(current_row)].value) + ' - ' + file_date_ext  # crystal ref
        portfolio_accounts['B{}'.format(portfolio_accounts_counter)].value = str(mapping_sheet2['D{}'.format(current_row)].value) # Port Acct
        portfolio_accounts['C{}'.format(portfolio_accounts_counter)].value = None # opening Units
        portfolio_accounts['D{}'.format(portfolio_accounts_counter)].value = float(mapping_sheet2['I{}'.format(current_row)].value) # Opening Nav
        portfolio_accounts['E{}'.format(portfolio_accounts_counter)].value = None # Opening FX Rate
        portfolio_accounts['F{}'.format(portfolio_accounts_counter)].value = mapping_sheet2['I{}'.format(current_row)].value * mapping_sheet2['J{}'.format(current_row)].value # Adj Opening Bal
        portfolio_accounts['G{}'.format(portfolio_accounts_counter)].value = investment_pool_funds[str(mapping_sheet2['D{}'.format(current_row)].value)] # Invst PnL
        portfolio_accounts['H{}'.format(portfolio_accounts_counter)].value = 0 # FX PnL
        portfolio_accounts['I{}'.format(portfolio_accounts_counter)].value = 0 # FX Hedge PnL
        portfolio_accounts['J{}'.format(portfolio_accounts_counter)].value = 0 # Interest income
        portfolio_accounts['K{}'.format(portfolio_accounts_counter)].value = 0 # int expense
        portfolio_accounts['L{}'.format(portfolio_accounts_counter)].value = 0 # admin fee
        portfolio_accounts['M{}'.format(portfolio_accounts_counter)].value = 0 # custody fee
        portfolio_accounts['N{}'.format(portfolio_accounts_counter)].value = 0 # misc expense
        portfolio_accounts['O{}'.format(portfolio_accounts_counter)].value = 0 # acct expenses
        portfolio_accounts['P{}'.format(portfolio_accounts_counter)].value = 0 # management fee
        portfolio_accounts['Q{}'.format(portfolio_accounts_counter)].value = 0 # perf fee
        portfolio_accounts['R{}'.format(portfolio_accounts_counter)].value = 0 # acct fees
        portfolio_accounts['S{}'.format(portfolio_accounts_counter)].value = portfolio_accounts['F{}'.format(portfolio_accounts_counter)].value + portfolio_accounts['G{}'.format(portfolio_accounts_counter)].value # closing bal
        portfolio_accounts['T{}'.format(portfolio_accounts_counter)].value = mapping_sheet2['J{}'.format(current_row)].value # closing units from mapping file
        portfolio_accounts['U{}'.format(portfolio_accounts_counter)].value = portfolio_accounts['S{}'.format(portfolio_accounts_counter)].value / portfolio_accounts['T{}'.format(portfolio_accounts_counter)].value # closing nav
        portfolio_accounts['V{}'.format(portfolio_accounts_counter)].value = 1

        # rewrite for mapping file edit
        mapping_sheet2['I{}'.format(current_row)].value = portfolio_accounts['U{}'.format(portfolio_accounts_counter)].value # multi
        mapping_sheet2['G{}'.format(current_row)].value = file_date_formatted # updated date
        mapping_sheet2['G{}'.format(current_row)].number_format = numbers.FORMAT_DATE_XLSX14

        portfolio_accounts_counter += 1

    ### Multi Portfolio
    ##portfolio_accounts['A{}'.format(portfolio_accounts_counter)].value = 'Franklin K2 Alternative Strategies Fund - ' + file_date_ext # crystal ref
    ##portfolio_accounts['B{}'.format(portfolio_accounts_counter)].value = 'Franklin K2 Alternative Strategies Fund Investment Pool' # Port Acct
    ##portfolio_accounts['C{}'.format(portfolio_accounts_counter)].value = None # opening Units
    ##portfolio_accounts['D{}'.format(portfolio_accounts_counter)].value = mapping_sheet2['I48'].value # Opening Nav
    ##portfolio_accounts['E{}'.format(portfolio_accounts_counter)].value = None # Opening FX Rate
    ##portfolio_accounts['F{}'.format(portfolio_accounts_counter)].value = mapping_sheet2['I48'].value * mapping_sheet2['J48'].value # Adj Opening Bal
    ##portfolio_accounts['G{}'.format(portfolio_accounts_counter)].value = adj_opening_multi # Invst PnL
    ##portfolio_accounts['H{}'.format(portfolio_accounts_counter)].value = 0 # FX PnL
    ##portfolio_accounts['I{}'.format(portfolio_accounts_counter)].value = 0 # FX Hedge PnL
    ##portfolio_accounts['J{}'.format(portfolio_accounts_counter)].value = 0 # Interest income
    ##portfolio_accounts['K{}'.format(portfolio_accounts_counter)].value = 0 # int expense
    ##portfolio_accounts['L{}'.format(portfolio_accounts_counter)].value = 0 # admin fee
    ##portfolio_accounts['M{}'.format(portfolio_accounts_counter)].value = 0 # custody fee
    ##portfolio_accounts['N{}'.format(portfolio_accounts_counter)].value = 0 # misc expense
    ##portfolio_accounts['O{}'.format(portfolio_accounts_counter)].value = 0 # acct expenses
    ##portfolio_accounts['P{}'.format(portfolio_accounts_counter)].value = 0 # management fee
    ##portfolio_accounts['Q{}'.format(portfolio_accounts_counter)].value = 0 # perf fee
    ##portfolio_accounts['R{}'.format(portfolio_accounts_counter)].value = 0 # acct fees
    ##portfolio_accounts['S{}'.format(portfolio_accounts_counter)].value = portfolio_accounts['F{}'.format(portfolio_accounts_counter)].value + portfolio_accounts['G{}'.format(portfolio_accounts_counter)].value # closing bal
    ##portfolio_accounts['T{}'.format(portfolio_accounts_counter)].value = mapping_sheet2['J48'].value # closing units from mapping file
    ##portfolio_accounts['U{}'.format(portfolio_accounts_counter)].value = portfolio_accounts['S{}'.format(portfolio_accounts_counter)].value / portfolio_accounts['T{}'.format(portfolio_accounts_counter)].value # closing nav
    ##portfolio_accounts['V{}'.format(portfolio_accounts_counter)].value = 1
    ##portfolio_accounts_counter += 1
    ##
    ### Long Short Portfolio
    ##portfolio_accounts['A{}'.format(portfolio_accounts_counter)].value = 'Franklin K2 Long Short Credit Fund - ' + file_date_ext # crystal ref
    ##portfolio_accounts['B{}'.format(portfolio_accounts_counter)].value = 'Franklin K2 Long Short Credit Fund Investment Pool' # Port Acct
    ##portfolio_accounts['C{}'.format(portfolio_accounts_counter)].value = None # opening Units
    ##portfolio_accounts['D{}'.format(portfolio_accounts_counter)].value = mapping_sheet2['I49'].value # Opening Nav
    ##portfolio_accounts['E{}'.format(portfolio_accounts_counter)].value = None # Opening FX Rate
    ##portfolio_accounts['F{}'.format(portfolio_accounts_counter)].value = mapping_sheet2['I49'].value * mapping_sheet2['J49'].value # Adj Opening Bal
    ##portfolio_accounts['G{}'.format(portfolio_accounts_counter)].value = adj_opening_ls # Invst PnL
    ##portfolio_accounts['H{}'.format(portfolio_accounts_counter)].value = 0 # FX PnL
    ##portfolio_accounts['I{}'.format(portfolio_accounts_counter)].value = 0 # FX Hedge PnL
    ##portfolio_accounts['J{}'.format(portfolio_accounts_counter)].value = 0 # Interest income
    ##portfolio_accounts['K{}'.format(portfolio_accounts_counter)].value = 0 # int expense
    ##portfolio_accounts['L{}'.format(portfolio_accounts_counter)].value = 0 # admin fee
    ##portfolio_accounts['M{}'.format(portfolio_accounts_counter)].value = 0 # custody fee
    ##portfolio_accounts['N{}'.format(portfolio_accounts_counter)].value = 0 # misc expense
    ##portfolio_accounts['O{}'.format(portfolio_accounts_counter)].value = 0 # acct expenses
    ##portfolio_accounts['P{}'.format(portfolio_accounts_counter)].value = 0 # management fee
    ##portfolio_accounts['Q{}'.format(portfolio_accounts_counter)].value = 0 # perf fee
    ##portfolio_accounts['R{}'.format(portfolio_accounts_counter)].value = 0 # acct fees
    ##portfolio_accounts['S{}'.format(portfolio_accounts_counter)].value = portfolio_accounts['F{}'.format(portfolio_accounts_counter)].value + portfolio_accounts['G{}'.format(portfolio_accounts_counter)].value # closing bal
    ##portfolio_accounts['T{}'.format(portfolio_accounts_counter)].value = mapping_sheet2['J49'].value # closing units from mapping file
    ##portfolio_accounts['U{}'.format(portfolio_accounts_counter)].value = portfolio_accounts['S{}'.format(portfolio_accounts_counter)].value / portfolio_accounts['T{}'.format(portfolio_accounts_counter)].value # closing nav
    ##portfolio_accounts['V{}'.format(portfolio_accounts_counter)].value = 1
    ##
    ##portfolio_crystal.delete_rows(portfolio_crystal.max_row + 1, 50)
    ##portfolio_positions.delete_rows(portfolio_positions.max_row + 1, 50)
    ##portfolio_accounts.delete_rows(portfolio_accounts.max_row + 1, 50)

    #--------------------------------------------------------------------------------------------------

    # transaction file start
    if (os=='Windows'):
        transaction_file = r'G:/Shared drives/Operations/K2/BNY - Daily/Transactions Template.xlsx'
    else:
        transaction_file = f'{folder_path}/templates/Transactions Template.xlsx'

    transaction = openpyxl.load_workbook(transaction_file)

    # open transaction worksheet
    transaction_sheet = transaction['Transactions']

    # start working on transactions
    UID = mapping_sheet3['B1'].value

    transaction_date = 'BNY Trx ' + file_date_formatted.strftime('%Y%m %d') + ' 0' + str(UID)

    # split for equity and investments

    # clean out sheet
    transaction_sheet.delete_rows(2, 1000)
                                            
    # start processing transaction sheet
    any_transactions = 0
    transaction_counter = 2

    for row in range(3, text_data_df_maxrow + 1):
        for acctnum in range(2, mapping_sheet2.max_row):
            if text_to_data_sheet['D{}'.format(row)].value == mapping_sheet2['D{}'.format(acctnum)].value:
                if mapping_sheet2['A{}'.format(acctnum)].value != None and mapping_sheet2['D{}'.format(acctnum)].value != None and mapping_sheet2['H{}'.format(acctnum)].value == True:
                    if text_to_data_sheet['H{}'.format(row)].value - mapping_sheet2['J{}'.format(acctnum)].value != 0:
                        transaction_sheet['A{}'.format(transaction_counter)].value = transaction_date # trans UID
                        UID += 1
                        transaction_date = 'BNY Trx ' + file_date_formatted.strftime('%Y%m %d') + ' 0' + str(UID)
                        transaction_sheet['B{}'.format(transaction_counter)].value = mapping_sheet2['E{}'.format(acctnum)].value # invstor acct long name
                        transaction_sheet['C{}'.format(transaction_counter)].value = mapping_sheet2['D{}'.format(acctnum)].value # invstment acct long name
                        transaction_sheet['D{}'.format(transaction_counter)].value = file_date_formatted # transaction date
                        transaction_sheet['D{}'.format(transaction_counter)].number_format = numbers.FORMAT_DATE_XLSX14
                        transaction_sheet['E{}'.format(transaction_counter)].value = file_date_formatted # custody date
                        transaction_sheet['E{}'.format(transaction_counter)].number_format = numbers.FORMAT_DATE_XLSX14
                        if text_to_data_sheet['H{}'.format(row)].value - mapping_sheet2['J{}'.format(acctnum)].value > 0:
                            transaction_sheet['F{}'.format(transaction_counter)].value = 'Buy' # transaction type
                            transaction_sheet['G{}'.format(transaction_counter)].value = abs(text_to_data_sheet['H{}'.format(row)].value - mapping_sheet2['J{}'.format(acctnum)].value) #units
                        else:
                            transaction_sheet['F{}'.format(transaction_counter)].value = 'Sell' # transaction type
                            transaction_sheet['G{}'.format(transaction_counter)].value = abs(text_to_data_sheet['H{}'.format(row)].value - mapping_sheet2['J{}'.format(acctnum)].value) #units
                        transaction_sheet['H{}'.format(transaction_counter)].value = mapping_sheet2['O{}'.format(acctnum)].value # NAV
                        transaction_sheet['I{}'.format(transaction_counter)].value = transaction_sheet['G{}'.format(transaction_counter)].value * transaction_sheet['H{}'.format(transaction_counter)].value # local invstment acct
                        transaction_sheet['J{}'.format(transaction_counter)].value = 1 # exchange rate
                        transaction_sheet['K{}'.format(transaction_counter)].value = transaction_sheet['I{}'.format(transaction_counter)].value # investor amt
                        transaction_sheet['L{}'.format(transaction_counter)].value = 0 # equalization
                        transaction_sheet['M{}'.format(transaction_counter)].value = 'Settled' # settled
                        transaction_counter += 1
                        any_transactions += 1

    transaction_sheet.delete_rows(transaction_sheet.max_row + 1, 50)

    if any_transactions > 0:
        transaction.save(transaction_file)
        if (os=='Windows'):
            transaction_dest = r'G:/Shared drives/Operations/K2/BNY - Daily/' + file_date_ext + ' Equity - Transactions K2 BNY TD ' + todayStr + '.xlsx'
        else:
            transaction_dest = f'{folder_path}/output/' + file_date_ext + ' Equity - Transactions K2 BNY TD ' + todayStr + '.xlsx'
        shutil.copyfile(transaction_file, transaction_dest) # create copied file
        print("Equity Transaction file created.\n")
    else:
        print("There are no Equity transactions.\n")

    #equity
    # clean out sheet
    transaction_sheet.delete_rows(2, 1000)
                                            
    # start processing transaction sheet
    any_transactions = 0
    transaction_counter = 2

    for row in range(3, text_data_df_maxrow + 1):
        for acctnum in range(2, mapping_sheet2.max_row):
            if text_to_data_sheet['D{}'.format(row)].value == mapping_sheet2['D{}'.format(acctnum)].value:
                if mapping_sheet2['A{}'.format(acctnum)].value != None and mapping_sheet2['D{}'.format(acctnum)].value != None and mapping_sheet2['H{}'.format(acctnum)].value == False:
                    if text_to_data_sheet['H{}'.format(row)].value - mapping_sheet2['J{}'.format(acctnum)].value != 0:
                        transaction_sheet['A{}'.format(transaction_counter)].value = transaction_date # trans UID
                        UID += 1
                        transaction_date = 'BNY Trx ' + file_date_formatted.strftime('%Y%m %d') + ' 0' + str(UID)
                        transaction_sheet['B{}'.format(transaction_counter)].value = mapping_sheet2['E{}'.format(acctnum)].value # invstor acct long name
                        transaction_sheet['C{}'.format(transaction_counter)].value = mapping_sheet2['D{}'.format(acctnum)].value # invstment acct long name
                        transaction_sheet['D{}'.format(transaction_counter)].value = file_date_formatted # transaction date
                        transaction_sheet['D{}'.format(transaction_counter)].number_format = numbers.FORMAT_DATE_XLSX14
                        transaction_sheet['E{}'.format(transaction_counter)].value = file_date_formatted # custody date
                        transaction_sheet['E{}'.format(transaction_counter)].number_format = numbers.FORMAT_DATE_XLSX14
                        if text_to_data_sheet['H{}'.format(row)].value - mapping_sheet2['J{}'.format(acctnum)].value > 0:
                            transaction_sheet['F{}'.format(transaction_counter)].value = 'Buy' # transaction type
                            transaction_sheet['G{}'.format(transaction_counter)].value = abs(text_to_data_sheet['H{}'.format(row)].value - mapping_sheet2['J{}'.format(acctnum)].value) #units
                        else:
                            transaction_sheet['F{}'.format(transaction_counter)].value = 'Sell' # transaction type
                            transaction_sheet['G{}'.format(transaction_counter)].value = abs(text_to_data_sheet['H{}'.format(row)].value - mapping_sheet2['J{}'.format(acctnum)].value) #units
                        transaction_sheet['H{}'.format(transaction_counter)].value = mapping_sheet2['I{}'.format(acctnum)].value # NAV
                        transaction_sheet['I{}'.format(transaction_counter)].value = transaction_sheet['G{}'.format(transaction_counter)].value * transaction_sheet['H{}'.format(transaction_counter)].value # local invstment acct
                        transaction_sheet['J{}'.format(transaction_counter)].value = 1 # exchange rate
                        transaction_sheet['K{}'.format(transaction_counter)].value = transaction_sheet['I{}'.format(transaction_counter)].value # investor amt
                        transaction_sheet['L{}'.format(transaction_counter)].value = 0 # equalization
                        transaction_sheet['M{}'.format(transaction_counter)].value = 'Settled' # settled
                        transaction_counter += 1
                        any_transactions += 1

    transaction_sheet.delete_rows(transaction_sheet.max_row + 1, 50)

    if any_transactions > 0:
        transaction.save(transaction_file)
        if (os=='Windows'):
            transaction_dest = r'G:/Shared drives/Operations/K2/BNY - Daily/' + file_date_ext + ' Investments - Transactions K2 BNY TD ' + todayStr + '.xlsx'
        else: 
            transaction_dest = f'{folder_path}/output/' + file_date_ext + ' Investments - Transactions K2 BNY TD ' + todayStr + '.xlsx'

        shutil.copyfile(transaction_file, transaction_dest) # create copied file
        print("Investments Transaction file created.\n")
    else:
        print("There are no Investments transactions.\n")

    # transaction file

    #--------------------------------------------------------------------------------------------------

    # mapping file edit start

    # have to rewrite over date (G)
    # Closing NAV (I)
    # Closing Units (J)

    update_counter = 0

    for row in range(3, text_data_df_maxrow):
        for acctnum in range(2, mapping_sheet2.max_row):
            if text_to_data_sheet['D{}'.format(row)].value == mapping_sheet2['D{}'.format(acctnum)].value and mapping_sheet2['D{}'.format(acctnum)].value != None:
                mapping_sheet2['G{}'.format(acctnum)].value = file_date_formatted # updated date
                mapping_sheet2['G{}'.format(acctnum)].number_format = numbers.FORMAT_DATE_XLSX14
                mapping_sheet2['I{}'.format(acctnum)].value = text_to_data_sheet['G{}'.format(row)].value # update NAV
                mapping_sheet2['J{}'.format(acctnum)].value = text_to_data_sheet['H{}'.format(row)].value # update oustanding shares
                update_counter += 1

    mapping_sheet3['B1'].value = 1
    mapping_sheet3['B2'].value = file_date_formatted
    mapping_sheet3['B2'].number_format = numbers.FORMAT_DATE_XLSX14

    print('There have been ' + str(update_counter) + ' updates to the mapping file.')

    read_mapping.save(mapping)

    # save and copy all files
    valuation.save(valuation_file)
    if (os=='Windows'):
        valuation_dest = r'G:/Shared drives/Operations/K2/BNY - Daily/' + file_date_ext + ' Valuation Data - Account Performance K2 BNY TD ' + todayStr + '.xlsx'
    else:
        valuation_dest = f'{folder_path}/output/' + file_date_ext + ' Valuation Data - Account Performance K2 BNY TD ' + todayStr + '.xlsx'
    shutil.copyfile(valuation_file, valuation_dest) # create copied file
    print("Valuation - NAV file created.\n")

    crystal.save(crystal_file)
    if (os=='Windows'):
        crystal_dest = r'G:/Shared drives/Operations/K2/BNY - Daily/' + file_date_ext + ' BNY Portfolio Crystallization ' + todayStr + '.xlsx'
    else:
        crystal_dest = f'{folder_path}/output/' + file_date_ext + ' BNY Portfolio Crystallization ' + todayStr + '.xlsx'
    shutil.copyfile(crystal_file, crystal_dest) # create copied file

    BNY_FX(mapping, crystal_dest)
    BNY_Close_Value_Update(mapping)

print("Crystallization file created.\n")
print("All files updated.\n")
