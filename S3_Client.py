import logging
from re import I
import boto3
from botocore.exceptions import ClientError
from botocore.config import Config
import platform as pt
import os
import sys
import io
import re
import pandas as pd
import datetime as dt
import msoffcrypto as mso
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border, NamedStyle, numbers
from datetime import datetime
import xlrd
import numpy as np
import json
#import s3fs


sys.path.insert(1,'/usr/local/share/python/lib') 
from EsmProperties import EsmProperties
from SQLProcessor import SQLProcessor

class S3_Client:
    def __init__(self, name='Linux',source=None,platform='AWS_DEV',debug=False,db_conn_str=None,db_uid=None,db_passwd=None,verify_ssl=True,access_key=None,secret_key=None,sqlproc=None):
        self.debug = debug
        self.now_stamp = datetime.now().strftime("%Y%m%d:%H:%M:%S")
        self.now = datetime.now()
        self.name = name
        self.sqlproc = sqlproc
        self.db_conn_str=db_conn_str
        self.db_uid=db_uid
        self.db_passwd=db_passwd
        self.aws_bucket_path = None
        self.aws_bucket_name = None
        self.verify_ssl=verify_ssl
        self.os_type = None
        if self.name != 'aws':
            _prop=EsmProperties("EsmProperties")
            self.os_type=pt.system()
            if ('Windows' in self.os_type):
                _prop.set_properties_file("G:/Shared drives/Common/Anatoliy/conf/esmjsql.ini")
                self.download_folder="G:/Shared drives/Common/Anatoliy/Temp"
                if self.debug:
                    print ("S3_Client ==> OS TYPE: {} paltforrm: {}".format(self.os_type,platform))
            else: 
                _prop.set_properties_file("/usr/share/vidrio/conf/esmjsql.ini")
                self.download_folder="/usr/local/share/vidrio/FileDownloads/S3"
            self.aws_access_key_id=_prop.get_property(f'{platform}','AWS_ACCESS_KEY_ID')
            self.aws_secret_access_key=_prop.get_property(f'{platform}','AWS_SECRET_ACCESS_KEY')
            self.aws_bucket_name_init=_prop.get_property(f'{platform}','AWS_BUCKET_NAME')
            self.aws_bucket_name_vidrio=_prop.get_property(f'{platform}','AWS_BUCKET_NAME_VIDRIO')
            self.aws_bucket_lambdas_path=_prop.get_property(f'{platform}','AWS_BUCKET_PATH_LAMBDAS')
            self.aws_bucket_layers_path=_prop.get_property(f'{platform}','AWS_BUCKET_PATH_LAYERS')
            if source is None:
                self.aws_bucket_path_init=_prop.get_property(f'{platform}','AWS_BUCKET_PATH')
            elif source == 'PPATH':
                self.aws_bucket_path_init=_prop.get_property(f'{platform}','AWS_BUCKET_PATH_PPATH')
            elif source == 'VANBIEMA':
                self.aws_bucket_path_init=_prop.get_property(f'{platform}','AWS_BUCKET_PATH_VB')
            elif source == 'OPSAUTOMATION':
                self.aws_bucket_path_init=_prop.get_property(f'{platform}','AWS_BUCKET_PATH_OPSAUTOMATION')
            elif source == 'ESMAUTOMATION':
                self.aws_bucket_path_init=_prop.get_property(f'{platform}','AWS_BUCKET_PATH_ESMAUTOMATION')
            elif source == 'ESMSECMASTER':
                self.aws_bucket_path_init=_prop.get_property(f'{platform}','AWS_BUCKET_PATH_ESMSECMASTER')
            elif source == 'MARTINTEST':
                self.aws_bucket_path_init=_prop.get_property(f'{platform}','AWS_BUCKET_PATH_MARTINTEST')
            elif source == 'ESM':
                self.aws_bucket_path_init=_prop.get_property(f'{platform}', 'AWS_BUCKET_PATH_ESM')
            else:
                self.aws_bucket_path_init=_prop.get_property(f'{platform}','AWS_BUCKET_PATH')
            self.aws_bucket_region=_prop.get_property(f'{platform}','AWS_BUCKET_REGION')
            if self.debug:
                print ("S3_Client ==> BUCKET PATH: {} REGION: {}".format(self.aws_bucket_path_init,self.aws_bucket_region))
        else:
            if access_key is None:
                self.aws_access_key_id=os.environ['ACCESS_KEY']
            else:
                self.aws_access_key_id = access_key 
            if secret_key is None:
                self.aws_secret_access_key=os.environ['SECRET_KEY']
            else:
                self.aws_secret_access_key  = secret_key 
            if self.debug:
                print("S3_Client ===> ACCESS KEY{} SECRET KEY{} : ".format(access_key,secret_key))
            self.aws_bucket_name_init=os.environ['S3_BUCKET_NAME']
            self.aws_bucket_path_init=os.environ['S3_BUCKET_PATH']
            self.aws_bucket_region=os.environ['AWS_LAMBDA_REGION']
        
        if self.debug:
            print("S3_Client ===> BUCKET PATH: ",self.aws_bucket_path_init,self.aws_bucket_path,self.aws_bucket_region)
        
        self.config = Config(
                read_timeout=54000,
                connect_timeout=54000,
                retries={"max_attempts": 0}
        )
        #getting dynamic keys for os based apps from [AWS01_Identity_Access_Management] -- overwriting config file
        if self.name != 'aws':
            if platform == 'AWS_DEV':
                self.__aws_user = 'd001-dev'
                self.__db = 'VIDRIODEV'
            elif platform == 'AWS_TEST':
                self.__aws_user = 'd001-test'
                self.__db = 'VIDRIOTEST'
            elif platform == 'AWS_LIVE':
                self.__aws_user = 'd001-live'
                self.__db = 'VIDRIOPROD'

            (self.aws_access_key_id,self.aws_secret_access_key) = self.__get_access_key_pair()

        if self.debug:
            print("S3_Client ==> ACCESS KEY ID : {} SECRET KEY: {}".format(self.aws_access_key_id,self.aws_secret_access_key))

        self.s3 = boto3.client('s3',aws_access_key_id=self.aws_access_key_id, aws_secret_access_key=self.aws_secret_access_key,region_name=self.aws_bucket_region,verify=self.verify_ssl,config=self.config)
        self.lambda_client = boto3.client('lambda',aws_access_key_id=self.aws_access_key_id, aws_secret_access_key=self.aws_secret_access_key,region_name=self.aws_bucket_region,verify=self.verify_ssl,config=self.config)
        self.s3_r = boto3.resource('s3',aws_access_key_id=self.aws_access_key_id, aws_secret_access_key=self.aws_secret_access_key,region_name=self.aws_bucket_region,verify=self.verify_ssl)

        self.df = None

    def __get_access_key_pair(self):
        processor = SQLProcessor(driver='pyodbc', db=self.__db)
        sp_name='{call [dbo].[AWS01_Retrieve_IAM_Users]}'
        sp_args = ()
        res = processor.execute_sp(sp_name=sp_name,sp_args=sp_args,get_result=True)
        keys = [x for x in res if x['User_Name']==self.__aws_user]
        return (keys[0]['Aws_Access_Key_Id'],keys[0]['Aws_Secret_Access_Key'])

    def generate_xml(self, output_dir = 'output_consolidated'):
        if self.sqlproc is None:
            processor = SQLProcessor('pyodbc', 'ESMPROD',connection_string=self.db_conn_str,uid=self.db_uid, password=self.db_passwd)
        else:
            processor = self.sqlproc

        sp_name = '{call [VIDRIO_INPUTS$Excel_to_xml_col_names]}'  

        # The following basically replaces names with invalid xml characters to valid characters
        # Data from database
        column_names = processor.execute_sp(sp_name=sp_name, sp_args=(), get_result=True)

        names_dict = {}
        for row in column_names:
            if row['template_sheet'] in names_dict.keys():
                names_dict[row['template_sheet']][row['template_col_xl']] = (row['template_col_xml'])
            else:
                names_dict[row['template_sheet']] = {row['template_col_xl'] : row['template_col_xml']}
        #print("NAMES:", names_dict)

        #iterate through files in output_consolidated
        files = self.find_files(sub_folder= self.aws_bucket_path + '/' + output_dir)
        for file in files:
            if file.split('.')[-1] != 'xlsx':
                continue

            if file.split('.')[-1] == 'xlsx':
                pattern_for_name = os.path.basename(file).split(' ')[-1].split('.')[0]
                pattern = os.path.basename(file).split('/')[-1]
                # self.pattern = pattern
            key = file.split('/')[-1]

            obj = self.s3.get_object(Bucket=self.aws_bucket_name, Key= self.aws_bucket_path + '/' + output_dir + '/' + key)
            # print('aws bucket name: ', self.aws_bucket_name)
            # print('aws bucket path: ', self.aws_bucket_path)
            # print('output xml: ', self.aws_bucket_name + '/' + self.aws_bucket_path + '/xml/K2_20221102-170813_Crystallization.xml')
            # output_xml_file = self.aws_bucket_name + '/' + self.aws_bucket_path + '/xml/K2_20221102-170813_Crystallization.xml'
            #print("OUTPUT XML FILE: ", output_xml_file)
            df_content = obj['Body'].read()

            number_of_sheets = len(pd.ExcelFile(df_content).sheet_names)
            sheet_names = pd.ExcelFile(df_content).sheet_names
            print('sheet names:', sheet_names)

            # #skip the first sheet and last sheet (Instructions and Lookup)

            for i in range(1,number_of_sheets - 1):
            # for sheet name in the file
                dataframe = pd.read_excel(df_content, sheet_name=i)
                #skip sheet if it is empty
                if dataframe.empty:
                    continue
                sheet_name = sheet_names[i]
                if self.debug:
                    print('S3_Client ===> sheet name: ', sheet_name)

                #row name variable
                row_name = None
                root_name = None
                for row in column_names:
                    # print#(row)
                    if(sheet_name == row['xl_sheet']):
                        row_name = row['template_sheet']
                        root_name = row['template_name']
                        #self.root_name is to pass the name to a variable outside this scope
                        self.root_name = root_name
                        # root_name = '<'+ root_name + '>'
                        if self.debug:
                            print('S3_Client ===> root name: ', root_name)
                        break
            
                for header in dataframe.columns:
                    for row in names_dict.values():
                        if header in row:
                            dataframe.rename(columns={header : row[header]}, inplace= True)
                            continue
                
                #obj_name variable
                output_xml_file = self.aws_bucket_name + '/' + self.aws_bucket_path + '/temp/' + file.split('/')[-1].split('.')[0] + '_' + sheet_name + '.xml'
                output_xml_file = output_xml_file.replace(' ', '_')
                obj_name = output_xml_file.split('/')[-1]
                print('obj name: ', obj_name)
                print('output xml file :', output_xml_file)

                print('written obj name: ', obj_name)
                self.df_to_xml(df = dataframe, obj_name= obj_name, root_name= root_name, row_name=row_name, load_to='temp', columns=list(dataframe.columns), index=False)    
            #remove from output_consolidated
            self.copy_file(key,dest_file=key,load_from='output_consolidated',load_to='processed',rm_src = True)

    def consolidate_xml(self):

            # xml sheets are in temp folder at this point, combine to one xml file and transer to xml folder
        # xl_root = str('<' + self.root_name + '>')
        # xl_root_end = xl_root.replace("<", "</")
        patterns = ('Crystallization', 'EquityTransactions', 'InvestmentTransactions', 'Pricing', 'FX_Forward')
        
        temp_files = self.find_files(sub_folder=self.aws_bucket_path + '/temp')
        split_files = []
        count = 1

        def _helper_last_file_merge(split_files, pattern, pattern_for_name):
            
            merged_content = ''
            f_cnt = 0
            for file in split_files:
                f = os.path.basename(file)
                if self.debug:
                    print('S3_Client ===> FILE NAME: ', f)
                # print("===================================================")
                content = str(self.read_text_file(f, load_from = 'temp', data_type= 'ascii'))
                content = content.replace("<?xml version='1.0' encoding='utf-8'?>\n","")
                lines = content.split("\n")
                xl_root = lines[0]
                # print('xl root: ', xl_root)
                xl_root_end = xl_root.replace("<", "</")
                # print("FILE COUNT: ", f_cnt)
                line_fixed = None

                if f_cnt == 0:
                        
                        merged_file = str(f).split(pattern_for_name)[0] + pattern_for_name + ".xml"
                        if self.debug:
                            print('S3_Client ===> XML FILE NAME: ', merged_file)
                        pattern_for_name = pattern

                for line in lines:

                    if ' 00:00:00' in line:
                        line_fixed = line.replace(" 00:00:00", "T00:00:00")
                                    
                    if line_fixed is not None:
                        line = line_fixed
                        line_fixed = None
                    
                    
                        
                    if f_cnt == 0 and xl_root_end not in line:
                        # print(f_cnt, 'ADDED LINE: ', line)
                        merged_content += line + "\n"

                    elif f_cnt > 0 and f_cnt < len(split_files)-1 and xl_root not in line and xl_root_end not in line:
                        merged_content += line + "\n"

                    # elif f_cnt == len(split_files)-1 and xl_root not in line:
                    #     merged_content += line + "\n"

                    elif f_cnt == len(split_files)-1 and xl_root not in line:
                        merged_content += line + '\n'

                # print("COUNT: ", f_cnt)
                # print("Num of files: ", len(split_files))
                # print("===================================================")
                f_cnt += 1
            self.write_to_ascii_file(merged_content, merged_file, load_to= 'xml')
        
        for file in temp_files:
            
            if self.debug:
                print('S3_Client ===> Count: ', count)
                print("S3_Client ===> FILE: ", file)
            name = file.split('/')[-1]
            if self.debug:
                print('S3_Client ===> Name: ',name)
            identifier = name.split('_')[1]
            # print('Identifier: ', identifier)
            pattern = None
            for pattern in patterns:
                if pattern in name:
                    
                    if split_files == []:
                        pattern_for_name = pattern
                        split_files.append(file)
                        f = os.path.basename(file)
                        if self.debug:
                            print('S3_Client ===> FILE NAME: ', f)
                        # print("===================================================")
                        content = str(self.read_text_file(f, load_from = 'temp', data_type= 'ascii'))
                        content = content.replace("<?xml version='1.0' encoding='utf-8'?>\n","")
                        lines = content.split("\n")
                        xl_root = lines[0]
                        # print('xl root: ', xl_root)
                        xl_root_end = xl_root.replace("<", "</")
                        # print('xl root end: ', xl_root_end)
                        
                        # if there is only 1 file in the temp directory
                        if len(temp_files) == count:
                            _helper_last_file_merge(split_files, pattern, pattern_for_name)

                            # delete files in temp folder
                            for temp_file in split_files:

                                print('temp xml file : ', temp_file)
                                self.delete_file(os.path.basename(temp_file),rm_from='temp')
                                print("File deleted.")

                    elif split_files != [] and identifier in split_files[0] and pattern in split_files[0]:
                        split_files.append(file)

                        #last file
                        if len(temp_files) == count:
                            _helper_last_file_merge(split_files, pattern, pattern_for_name)
                            pattern_for_name = pattern
                        
                    
                    # elif this is a new file that does not go with the previous files
                    elif ((split_files != [] and pattern not in split_files[0] and not(pattern in split_files[0] and identifier not in split_files[0]))
                    or
                    (pattern in split_files[0] and identifier not in split_files[0])):
                        _helper_last_file_merge(split_files, pattern, pattern_for_name)
                        pattern_for_name = pattern
                        
                        # delete files in temp folder
                        for temp_file in split_files:

                            print('temp xml file : ', temp_file)
                            self.delete_file(os.path.basename(temp_file),rm_from='temp')
                            print("File deleted.")
                        if self.debug:
                            print('S3_Client ===> BEFORE CLEAR: ',split_files)
                        split_files.clear()
                        if self.debug:
                            print('S3_Client ===> AFTER CLEAR: ',split_files)

                        split_files.append(self.aws_bucket_path + '/temp/' + name)
                        pattern_for_name = pattern
                        if temp_files[-1] == split_files[0]:
                            _helper_last_file_merge(split_files, pattern, pattern_for_name)
                                    
                            # delete files in temp folder
                            for temp_file in split_files:
                                print('temp xml file : ', temp_file)
                                self.delete_file(os.path.basename(temp_file),rm_from='temp')
                                print("File deleted.",temp_file)
                    break
            count += 1
            # print(count)
            # catching the last file
            if self.debug:
                print("S3_Client ===> LEN: ", len(temp_files))

        #cleanup
        #for file in temp_files:
        #   print('temp xml file : ', temp_file)
        #    self.delete_file(temp_file,rm_from='temp')
        #    print("File deleted.",temp_file)
        self.archive_files(src_dir = 'temp' , target_dir = 'archive')
        

    def set_AWS_attributes(self,cloud_path,file_name):
        #match = re.match('(^[\S\w]+)/([\S\w]+)/([\S\w]+)$',cloud_path.replace("s3://",""))
        if "s3://" in cloud_path:
            b_path = os.path.dirname(cloud_path).replace("s3://","")
            b_file = os.path.basename(cloud_path).replace("s3://","")
            match = re.match('(^[^/]+)/(.+)$',b_path)
            self.aws_bucket_name=match.group(1)
            self.aws_bucket_path = match.group(2) 
            self.aws_bucket_file = b_file
        else:
            self.aws_bucket_path = os.path.dirname(cloud_path)
            self.aws_bucket_file = os.path.basename(cloud_path)
            self.aws_bucket_real_file = file_name

    def reset_AWS_attributes(self):
        self.aws_bucket_name=self.aws_bucket_name_init
        self.aws_bucket_path=self.aws_bucket_path_init
    
    def normalize_AWS_attributes(self):
        if self.aws_bucket_name is None:
            self.aws_bucket_name = self.aws_bucket_name_init
        if self.aws_bucket_path is None:
            self.aws_bucket_path = self.aws_bucket_path_init

    def upload_file(self,file_name=None, content=None,object_name=None,load_to=None,lambda_zip=False,layers_zip=False):
        """Upload a file to an S3 bucket

        :param file_name: File to upload
        :param bucket: Bucket to upload to
        :param object_name: S3 object name. If not specified then file_name is used
        :return: True if file was uploaded, else False
        """

        
        self.normalize_AWS_attributes()

        # If S3 object_name was not specified, use file_name
        if not lambda_zip and not layers_zip:
            if object_name is None and file_name is not None:
                if load_to is None:
                    object_name = self.aws_bucket_path+"/"+os.path.basename(file_name)
                else:
                    object_name = self.aws_bucket_path+"/"+load_to+"/"+os.path.basename(file_name)
            else:
                if load_to is None:
                    object_name = self.aws_bucket_path+"/"+object_name
                else:
                    object_name = self.aws_bucket_path+"/"+load_to+"/"+object_name
        elif layers_zip:
            print("upload_file ==> Uploading Layers")
            if object_name is None and file_name is not None:
                if load_to is None:
                    object_name = self.aws_bucket_layers_path+"/"+os.path.basename(file_name)
                else:
                    object_name = self.aws_bucket_layers_path+"/"+load_to+"/"+os.path.basename(file_name)
            else:
                if load_to is None:
                    object_name = self.aws_bucket_layers_path+"/"+object_name
                else:
                    object_name = self.aws_bucket_layers_path+"/"+load_to+"/"+object_name
        else:
            if object_name is None and file_name is not None:
                if load_to is None:
                    object_name = self.aws_bucket_lambdas_path+"/"+os.path.basename(file_name)
                else:
                    object_name = self.aws_bucket_lambdas_path+"/"+load_to+"/"+os.path.basename(file_name)
            else:
                if load_to is None:
                    object_name = self.aws_bucket_lambdas_path+"/"+object_name
                else:
                    object_name = self.aws_bucket_lambdas_path+"/"+load_to+"/"+object_name

        if self.debug:
            print("S3_Client ===> BUCKET: ",self.aws_bucket_name," UPLOAD OBJECT NAME:",object_name, " FILE NAME: ", file_name, " Content: " , ('NONE' if content is None else 'CONTENT'))

        if content is None and file_name is not None:
            with open(file_name, 'rb') as f:
                body=f.read()
        else:
            body = content

        # Upload the file
        try:
            response = self.s3.put_object(Body=body, Bucket=self.aws_bucket_name, Key=object_name)
            if self.debug:
                print("S3_Client ===> UPLOAD RESPONSE: " , response , " FOR THE KEY: ",object_name , " AND BUCKET: ",self.aws_bucket_name)
        except ClientError as e:
            logging.error(e)
            return False
        return True

    def upload_dataframe(self,df=None, object_name=None,load_to=None,object_type='xlsx'):
        """Upload a data frame to an S3 bucket

        :param file_name: File to upload
        :param bucket: Bucket to upload to
        :param object_name: S3 object name. If not specified then file_name is used
        :return: True if file was uploaded, else False
        """

        self.normalize_AWS_attributes()

        # If S3 object_name was not specified, use file_name
        if object_name is None:
            if load_to is None:
                object_name = self.aws_bucket_path+"/"+os.path.basename(object_name)
            else:
                object_name = self.aws_bucket_path+"/"+load_to+"/"+os.path.basename(object_name)
        else:
            if load_to is None:
                object_name = self.aws_bucket_path+"/"+object_name
            else:
                object_name = self.aws_bucket_path+"/"+load_to+"/"+object_name

        if df is not None:
            if self.debug:
                print("S3_Client ===> DATAFREME TO UPLOAD:",df)
            with io.BytesIO() as output:
                if object_type!='json':
                    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                        if 'dict' not in str(type(df)): #one tab excel
                            df.to_excel(writer,index=False)
                        else: #multiple tabs excel
                            for tab_name in df.keys():
                                df_tab = df[tab_name] 
                                if self.debug:
                                    print("ADDING TAB :",tab_name, " DF: ",str(type(df_tab)))
                                df_tab.to_excel(writer,index=False,sheet_name=tab_name)
                        writer.close()
                        output.seek(0)
                        data = output.getvalue()
                        #print("DATA:",data)
                else:
                    df.to_json(output,orient='records')
                    output.seek(0)
                     

                # Upload the data frame
                try:
                    response = self.s3.put_object(Body=data, Bucket=self.aws_bucket_name, Key=object_name)
                    if self.debug:
                        print("S3_Client ===> UPLOAD RESPONSE: " , response , "  FOR OBJ: ",object_name)
                except ClientError as e:
                    logging.error(e)
                    return False
            return True
        else:
            return False

    def write_to_ascii_file(self,content, object_name,load_to=None):
        """Upload a file to an S3 bucket

        :param file_name: File to upload
        :param bucket: Bucket to upload to
        :param object_name: S3 object name. If not specified then file_name is used
        :return: True if file was uploaded, else False
        """

        self.normalize_AWS_attributes()

        # If S3 object_name was not specified, use file_name
        if load_to is None:
            object_name = self.aws_bucket_path+"/"+object_name
        else:
            object_name = self.aws_bucket_path+"/"+load_to+"/"+object_name
        # Upload the file
        if self.debug:
            print("S3_Client ===> PUT OBJECT NAME:",object_name)
        try:
            response = self.s3.put_object(Body=content, Bucket=self.aws_bucket_name, Key=object_name) 
            if self.debug:
                print("S3_Client ===> UPLOAD RESPONSE: " , response)
        except ClientError as e:
            logging.error(e)
            return False
        return True


    #ascii data by default
    def download_file(self,object_name,load_from=None,download_to=None,data_type=None,write_to_file=False,write_to_df=False,read_tab='unspecified',type='xlsx',x_path='.//BaseCCY',name_spaces=None,delimiter='|',header_row=1,skiprows=None,password=None,nrows=None):
        
        if self.debug:
            print("S3_Client ===> download_file ===> ",object_name)
        self.normalize_AWS_attributes()
        try:
            if load_from is None:
                in_key = self.aws_bucket_path+"/"+object_name
            else:
                in_key = self.aws_bucket_path+"/"+load_from+"/"+object_name
            if self.debug:
                print("S3_Client ===> BUCKET:",self.aws_bucket_name,"DOWNLOAD KEY:",in_key)
            if load_from is None:
                response =  self.s3.get_object(Bucket=self.aws_bucket_name, Key=in_key)
            else:
                response =  self.s3.get_object(Bucket=self.aws_bucket_name, Key=in_key)
            fileObj = response['Body']

            if data_type is None or (data_type == 'ascii' and type != 'csv'):
                content = io.BytesIO(fileObj.read()).read().decode('UTF-8')
            else:
                content = io.BytesIO(fileObj.read())

            

            if write_to_file and (data_type is None or data_type == 'ascii'):
                if download_to is None:
                    with open(f'{self.download_folder}/{object_name}', 'w') as f:
                        f.write(content)
                else:
                    with open(f'{download_to}/{object_name.replace(":","_")}', 'w') as f:
                        f.write(content)

            elif write_to_file and data_type == 'bin':
                if download_to is None:
                    with open(f'{self.download_folder}/{object_name}', 'wb') as f:
                        f.write(content.read())
                else:
                    with open(f'{download_to}/{object_name}', 'wb') as f:
                        f.write(content.read())

            elif not write_to_file and write_to_df:
                if self.debug:
                    print("S3_Client ===> S3 download_file file content: ",content, "  READ TAB: " , read_tab, "  TYPE: " , type)
                if type == 'xml':
                    if x_path is not None and name_spaces is not None:
                        self.df=pd.read_xml(content,xpath=x_path,namespaces=name_spaces)
                    elif x_path is not None and name_spaces is None:
                        self.df=pd.read_xml(content,xpath=x_path)
                    else:
                        self.df=pd.read_xml(content)
                elif type == 'csv':
                    if self.debug:
                        print("S3_Client ===> S3 download_file file csv content: ",content)
                    self.df=pd.read_csv(content,delimiter=delimiter,header=header_row, skiprows=skiprows,nrows=nrows,encoding='utf8')
                else:
                    #if read_tab is not None:
                    if read_tab == 'unspecified':
                        self.df=pd.read_excel(content,skiprows=skiprows,nrows=nrows)
                    else:
                        print("S3_Client ===> S3 download_file file {} content for tab {}".format(type,read_tab))
                        self.df=pd.read_excel(content,sheet_name=read_tab,skiprows=skiprows,nrows=nrows)
                    #else:
                    #    self.df=pd.read_excel(content)
        except ClientError as e:
            logging.error(e)
            return None
        
        if not write_to_df:
            return content
        else:
            if self.debug:
                print("S3_Client ===> S3 download_file ====> ",self.df)
            return self.df

    def readPasswordProtectedXLFileFromS3(self,objectname, password, load_from=None,read_tab='unspecified',**kwargs):
        """ A function to read/load a password-protected Excel file without having to manually type it in Excel interface
        Args:
            objectPath (str): The path to the object in the S3 bucket.
            password (str): The password to the object.
            **kwargs: Any additional arguments to pass to the Pandas's read_excel method. E.g. header=1, sheet_name='testSheet'
        Returns:
        DataFrame: Pandas and Spark DataFrames.
        """
  
        try:
            objectPath = self.aws_bucket_path+"/"+objectname if load_from is None else self.aws_bucket_path+"/"+load_from+'/'+objectname
            # mount s3 like local fs
            #self.s3 = s3fs.S3FileSystem (anon=False)

            if self.debug:
                print("S3_Client ===> readPasswordProtectedXLFileFromS3 ==> BUCKET NAME:", self.aws_bucket_name,"OBJECT PATH:",objectPath)
            # create an in-memory ByteIO object 
            decrypted_wb = io.BytesIO()

            response =  self.s3.get_object(Bucket=self.aws_bucket_name, Key=objectPath)
            fileObj = response['Body']
            content = io.BytesIO(fileObj.read())


            if self.name != 'aws' and self.os_type == 'Windows':
                with open ('C:/Temp/encr_excel_test.xlsx','wb') as f:
                    f.write(content.read())
            
                decrypted = io.BytesIO()
                with open('C:/Temp/encr_excel_test.xlsx', "rb") as f:
                    print ("S3_Client ===> FILE OBJ:",f)
                    data = mso.OfficeFile(f)
                    # Default passwords for encrypted excel sheets
                    # Add your password here if it differs than the default
                    data.load_key(password=password)
                    data.decrypt(decrypted)
                    df = pd.read_excel(decrypted)
            
                #os._exit(0)

            if self.debug:
                print("S3_Client ===> DECRYPTED CONTENT:",type(content))
            with  io.BufferedReader (content) as xlscontent:
    
                if self.debug:
                    print("S3_Client ===> CONTENT:",xlscontent)
                # open the protected file
                office_file = mso.OfficeFile(xlscontent)
                if self.debug:
                    print("S3_Client ===> OFFICE FILE:",office_file,'PASSWORD: ',password)
    
                # provide the password
                office_file.load_key(password=password)
    
                decrypted_wb = io.BytesIO()
                # decrypt and write to output file
                office_file.decrypt(decrypted_wb)

                if read_tab == 'unspecified':
                    self.df=pd.read_excel(decrypted_wb)
                else:
                    self.df=pd.read_excel(decrypted_wb,sheet_name=read_tab)
                return self.df
        except Exception as e:
            logging.error(e)

    def read_text_file(self,object_name,data_type='ascii',load_from=None,write_to_file=False,write_to_df=False,read_tab=None,separator=',',header_row=0, nrows=None,error_bad_lines='warn',from_stream=False, skip_blank_lines=False,return_file_object=False):
        try:
            if load_from is None:
                in_key = self.aws_bucket_path+"/"+object_name
            else:
                in_key = self.aws_bucket_path+"/"+load_from+"/"+object_name
            if self.debug:
                print("S3_Client ===> read_text_file Key ==>",in_key)
            if load_from is None:
                response =  self.s3.get_object(Bucket=self.aws_bucket_name, Key=in_key)
            else:
                response =  self.s3.get_object(Bucket=self.aws_bucket_name, Key=in_key)
            fileObj = response['Body']
            if data_type is None or data_type == 'ascii':
                content = io.BytesIO(fileObj.read()).read().decode('UTF-8')
            else:
                content = io.BytesIO(fileObj.read())
            #content extraction is finished
            if write_to_file and (data_type is None or data_type == 'ascii'):
                with open(f'{self.download_folder}/{object_name}', 'w') as f:
                    f.write(content)
            elif write_to_file and data_type == 'bin':
                with open(f'{self.download_folder}/{object_name}', 'wb') as f:
                    f.write(content.read())
            elif not write_to_file and write_to_df and data_type == 'bin':
                if read_tab is None:
                    self.df=pd.read_excel(content)
                else:
                    self.df=pd.read_excel(content,sheet_name=read_tab)
            elif not write_to_file and write_to_df and data_type == 'ascii':
                df = self.text_to_df(content=content,header_row=header_row,separator=separator, nrows=nrows,skip_rows=header_row,error_bad_lines=error_bad_lines,from_stream=from_stream, skip_blank_lines=skip_blank_lines)
        except ClientError as e:
            logging.error(e)
            return None

        if not write_to_df and not return_file_object:
            return content
        if not write_to_df and return_file_object:
            return fileObj
        else:
            return df

    
    def get_current_df(self):
        return self.df
    
    def get_init_bucket_path(self):
        return self.aws_bucket_path_init

    def df_to_excel(self,dfs,sheet_names,obj_name,load_to=None):
        if load_to is None:
            obj_name = self.aws_bucket_path+"/"+obj_name
        else:
            obj_name = self.aws_bucket_path+"/"+load_to+"/"+obj_name

        idx=0
        with io.BytesIO() as output:
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                for df in dfs:
                    df.to_excel(writer,sheet_name=sheet_names[idx],index=False)
                    idx += 1
            data = output.getvalue()

        self.s3.put_object(Key=obj_name, Bucket=self.aws_bucket_name, Body=data)

    def df_to_xml(self,df,obj_name,root_name,row_name,columns,index=False,load_to=None):
        if load_to is None:
            obj_name = self.aws_bucket_path+"/"+obj_name
        else:
            obj_name = self.aws_bucket_path+"/"+load_to+"/"+obj_name

        idx=0
        with io.BytesIO() as output:
            df.to_xml(output,root_name=root_name,row_name=row_name,attr_cols=columns,index=index)
            data = output.getvalue()
        self.s3.put_object(Key=obj_name, Bucket=self.aws_bucket_name, Body=data)
    

    def create_subfolder(self,subfolder):
        try:
            self.s3.put_object(Bucket=self.aws_bucket_name, Key=self.aws_bucket_path+'/'+subfolder)
        except ClientError as e:
            logging.error(e)
            return False
        return True

    def find_files(self,file_pattern=None,sub_folder=None,as_of=None,when=None):
        # method to get list of files from a bucket
        self.reset_AWS_attributes()
        files = list()
        bucket = self.s3_r.Bucket(self.aws_bucket_name)
        if self.aws_bucket_path in sub_folder:
            sub_folder = sub_folder.split("/")[-1]   
        if self.debug:
            print("S3_Client ===> find_files ==> Bucket: ",self.aws_bucket_name , "  Sub Folder: ", sub_folder)
        if sub_folder is None:
            for obj in bucket.objects.all():
                if file_pattern is None:
                    print(obj.key, obj.last_modified)
                    if (as_of is not None and when is None and obj.last_modified.date() == as_of.date()):
                        files.append(obj.key)
                    elif (as_of is not None and when == 'older' and obj.last_modified.date() < as_of.date()):
                        files.append(obj.key)
                    elif (as_of is not None and when == 'newer' and obj.last_modified.date() > as_of.date()):
                        files.append(obj.key)
                    elif (as_of is None):
                        files.append(obj.key)
                elif file_pattern in str(obj.key):
                    if (as_of is not None and when is None and obj.last_modified.date() == as_of.date()):
                        files.append(obj.key)
                    elif (as_of is not None and when == 'older' and obj.last_modified.date() < as_of.date()):
                        files.append(obj.key)
                    elif (as_of is not None and when == 'newer' and obj.last_modified.date() > as_of.date()):
                        files.append(obj.key)
                    elif (as_of is None):
                        files.append(obj.key)
        else:
            for obj in bucket.objects.filter(Prefix=self.aws_bucket_path+"/"+sub_folder+"/"):
                if self.debug:
                    print("S3_Client ===> find_files ==> obj",obj.key,obj.last_modified)
                if file_pattern is None and obj.key is not None and not str(obj.key).endswith("/"): 
                    if (as_of is not None and when is None and obj.last_modified.date() == as_of.date()):
                        files.append(obj.key)
                    elif (as_of is not None and when == 'older' and obj.last_modified.date() < as_of.date()):
                        files.append(obj.key)
                    elif (as_of is not None and when == 'newer' and obj.last_modified.date() > as_of.date()):
                        files.append(obj.key)
                    elif (as_of is None):
                        files.append(obj.key)

                elif file_pattern is not None and file_pattern in str(obj.key) and obj.key is not None:
                    if (as_of is not None and when is None and obj.last_modified.date() == as_of.date()):
                        files.append(obj.key)
                    elif (as_of is not None and when == 'older' and obj.last_modified.date() < as_of.date()):
                        files.append(obj.key)
                    elif (as_of is not None and when == 'newer' and obj.last_modified.date() > as_of.date()):
                        files.append(obj.key)
                    elif (as_of is None):
                        files.append(obj.key)
        if self.debug:
            print("find_files ==> ", files)
        return files

   

    def copy_file_direct(self,src_path,dest_file,src_file=None,load_to=None,rm_src=False,use_object_only=False):
        self.set_AWS_attributes(src_path,dest_file)
        src_obj_key = self.aws_bucket_path+"/"+self.aws_bucket_file
        print("copy_file_direct ==> key: ",src_obj_key)
        copy_source = {
            'Bucket': self.aws_bucket_name,
            'Key': src_obj_key
        }
        self.reset_AWS_attributes()
        try:
            if load_to is not None:
                dest_key = self.aws_bucket_path+"/"+load_to+"/"+dest_file
                self.s3_r.meta.client.copy(copy_source,self.aws_bucket_name,dest_key)
            else:
                dest_key = self.aws_bucket_path+"/"+dest_file
                self.s3_r.meta.client.copy(copy_source,self.aws_bucket_name,dest_key)
            if self.debug:
                print("S3_Client ===> COPY SOURCE: {} DEST BUCKET: {} DEST KEY: {} WITH RM SRC: {}".format(copy_source,self.aws_bucket_name,dest_key,rm_src))
            if rm_src:        
                response = self.delete_file(src_obj_key,use_obj_name_only=use_object_only)
                if self.debug:
                    print("S3_Client ===> DELETE RESPONSE: " , response , " FOR THE SRC OBJ: ",src_obj_key)
            return True
        except Exception as e:
            logging.error(e)
            return False
    

    def archive_files(self, src_dir = None, target_dir = None, pattern=None, skip_pattern=None):
        if pattern is None:
            files = self.find_files(sub_folder= self.aws_bucket_path + '/' + src_dir)
        else:
            files = self.find_files(sub_folder= self.aws_bucket_path + '/' + src_dir, file_pattern=pattern)
        print("Files in ", src_dir, " directory.")
        print(files)

        for path in files:
            file = str(path).split('/')[-1]
            if skip_pattern is None or skip_pattern not in file:
                self.copy_file(src_file=file, dest_file=file, load_from=src_dir, load_to=target_dir, rm_src=True, time_stamp= True)

    def copy_file(self,src_file, dest_file, load_from=None, load_to=None,rm_src=False, time_stamp = True):
        if load_from is None:
            obj_name_src = src_file
        else:
            obj_name_src = load_from+"/"+src_file

        if load_to is None:
            if time_stamp:
                obj_name_dest = f'{self.now_stamp}_{dest_file}'
            else:
                obj_name_dest = dest_file
        else:
            if time_stamp:
                obj_name_dest = load_to+"/"+f'{self.now_stamp}_{dest_file}'
            else:
                obj_name_dest = load_to + '/' + dest_file
        # copy file
        if self.debug:
            print("S3_Client ===> COPY THE SRC KEY: ",obj_name_src , " AND THE DEST KKEY: ",obj_name_dest)
        try:
            #response = self.s3.copy_object(CopySource=obj_name_src, Bucket=self.aws_bucket_name, Key=obj_name_dest,ACL='public-read')
            #response = self.s3_r.Object(self.aws_bucket_name, obj_name_dest).copy_from(CopySource=obj_name_src)
            content = self.download_file(object_name=obj_name_src,data_type='bin')
            response = self.upload_file(content=content,object_name=obj_name_dest)
            if self.debug:
                print("S3_Client ===> COPY RESPONSE: " , response , " FOR THE SRC KEY: ",obj_name_src , " AND THE DEST KKEY: ",obj_name_dest)
            if rm_src:
                response = self.delete_file(obj_name_src)
                if self.debug:
                    print("S3_Client ===> DELETE RESPONSE: " , response , " FOR THE SRC OBJ: ",obj_name_src)
        except Exception as e:
            logging.error(e)
            return False
        return True
    
    def get_file_from_cloud_path(self,cloud_path,file_name,load_to=None):
        self.set_AWS_attributes(cloud_path,file_name)
        if '.xml' in self.aws_bucket_real_file: 
            type='xml'
            data_type = 'ascii'
        elif '.xls' in self.aws_bucket_real_file or '.xlsx' in self.aws_bucket_real_file:
            type='xlsx'
            data_type = 'bin'
        else: # ascii for csv and txt
            type='csv'
            data_type = 'ascii'
        buf = self.download_file(self.aws_bucket_file,data_type=data_type,write_to_df=True,type=type,read_tab=None)
        #print(buf)
        if self.debug:
            print( "S3_Client ===> CLOUD PATH: ",self.aws_bucket_name,self.aws_bucket_file,self.aws_bucket_path,self.aws_bucket_real_file)
        if '.xml'  in self.aws_bucket_real_file or '.txt' in self.aws_bucket_real_file:
            content=self.read_text_file(self.aws_bucket_file)
            if self.debug:
                print ("S3_Client ===> ASCII CONTENT: ",content)
        else:
            content=self.read_text_file(self.aws_bucket_file,data_type='bin')
            if self.debug:
                print ("S3_Client ===> BIN CONTENT: ",content)
            
        

        self.reset_AWS_attributes()
        if self.debug:
            print("S3_Client ===> get_file_from_cloud_path ==> BEFORE UPLOAD: BUCKET NAME: " ,self.aws_bucket_name, "  BUCKET PATH: " ,self.aws_bucket_path, "   OBJ NAME: ", self.aws_bucket_real_file, " LOAD TO : ", load_to)

        self.upload_file(file_name=self.aws_bucket_real_file, content=content,object_name=self.aws_bucket_real_file,load_to=load_to)


    def delete_file(self,obj_name,rm_from=None,use_obj_name_only=False):

        if rm_from is None and not use_obj_name_only:
            obj_name = self.aws_bucket_path+"/"+obj_name
        elif rm_from is not None and not use_obj_name_only:
            obj_name = self.aws_bucket_path+"/"+rm_from+"/"+obj_name
        else: #just for readability
            obj_name = obj_name

        response = self.s3.delete_object(Bucket=self.aws_bucket_name,Key=obj_name)
        if self.debug:
            print("S3_Client ===>DELETE RESPONSE: " , response , " FOR THE KEY: ",obj_name , " AND BUCKET: ",self.aws_bucket_name)
        return response

    def set_bucket_name(self,bucket_name):
        self.aws_bucket_name=bucket_name 

    def set_bucket_path(self,bucket_path):
        self.aws_bucket_path = bucket_path

    def trigger_lambda(self,function_name=None,invocation_type='Event', lambda_payload=b'{}'):
        if function_name is not None:
            response = self.lambda_client.invoke(FunctionName=function_name,
                     InvocationType=invocation_type,
                     Payload=lambda_payload)
            if self.debug:
                print("trigger_lambda RESPONSE: ", response)
        else:
            print("No Function To Invoke!!")
        
        return response

    def _remove_num_format(self,buf):
        start = False
        end = False
        i=0
        ltr = list()
        for c in buf:
            if not start:
                if c == '"':
                    start=True
                    end = False
                else:
                    ltr.append(c)
            elif not end:
                if c != ',':
                    if c == '"' :
                        start = False
                        end = True
                    else:
                        ltr.append(c)
            
        buf = ''
        for c in ltr:
            if c!="\r":
                buf += c
        return buf
    
    def text_to_df (self,content,header_row,separator, nrows = None,from_stream=False,error_bad_lines='warn',skip_rows=0, skip_blank_lines=False):
        lines = content.split("\n")
        header = lines[header_row]
        h_cols = header.replace("\r","").split(separator)
        if self.debug:
            print("text_to_df ==> HEADER:",h_cols)
        data=list()
        k=0
        for line in lines:
            row = self._remove_num_format(line).split(separator)

            if len(row)<5 or k <= header_row:
                if self.debug:
                    print("text_to_df ==> INCOMPLETE ROW:",row,"  LENGTH:" , len(row),k)
            else:
                if self.debug:
                    print("text_to_df ==> FULL ROW:",row," LENGTH:", len(row),k)
                if nrows is not None and k >= nrows:
                    break
                data.append(row)
            k += 1
        #print("DATA: ",data)
        if not from_stream:
            df=pd.DataFrame(data, columns=h_cols)
            if self.debug and self.name != 'aws':
                with pd.ExcelWriter('C:/temp/text_to_df.xlsx',engine='xlsxwriter') as writer:
                    df.to_excel(writer,index=False) 
        elif from_stream:
            df = pd.read_csv(io.StringIO(content),on_bad_lines=error_bad_lines,skiprows=skip_rows, skip_blank_lines=skip_blank_lines)

        return df
    
    #Returns size of a bucket object in MB
    def get_object_size(self,in_key):
        obj = self.s3_r.ObjectSummary(self.aws_bucket_name, in_key)
        return round(obj.size/1024/1024,2)
    
    def update_function_env_vars(self,function_name):
        response = self.lambda_client.update_function_configuration(
            FunctionName=function_name,
            Environment={
                'Variables': {
                    'AWS_ENVIRONMENT_CODE': 'TEST',
                    'AWS_LAMBDA_REGION': 'eu-west-1',
                    'AWS_USER_NAME': 'd001-test',
                    'To': 'esm_users_ops@vidrio.com',
                    'CC': 'anatoliy.shor@vidrio.com',
                    'BCC': 'anatoliy.shor@vidrio.com',
                    'CONFIG_FOLDER':'config',
                    'INPUT_FOLDER': 'input',
                    'LOG_GROUP': 'ESM',
                    'S3_BUCKET_NAME': 'com.vidrio.lambdalayer-test',
                    'S3_BUCKET_PATH': 'DataFiles/OPSAutomation',
                    'TZ': 'US/Eastern',
                    'DBUSER': 'esm_sa',
                    'DBPASSWD': 'esm15vidrio'
                }
            }
        )
        return response

if __name__ == "__main__":
    
    drvr = S3_Client(name = 'Windows', source='ESMAUTOMATION', platform='AWS_LIVE',debug=True)
    drvr.reset_AWS_attributes()

    #payload=dict()
    #payload['proc_ident'] = 1098845
    #payload['two_files_monitor'] =  0
    #payload['entity_long_name'] = 'MW TOPS UCITS Fund'
    #payload['entity_id'] = '452484|6304|403386'
    #payload['process_as_of'] = '05/31/2024'
    #payload['process_period'] = '2024M05'
    #payload['vidrio_proc_template_uid'] = 124
    #payload['validator_uid'] = 138
    #payload['file_count'] = 1
    #response = drvr.trigger_lambda(function_name='ESM-DataAcquisition-Large',invocation_type='RequestResponse',lambda_payload=json.dumps(payload))
    #print(response)

    #cloud_path = 'WebDocumentStorage/1384947320246460710303157.enc'
    #cloud_path = 's3://com.vidrio.d001-test/WebDocumentStorage/1384947520244220715302716.enc'
    #file_name = 'Egerton.xlsx'
    #drvr.copy_file_direct(cloud_path,file_name,load_to='input')
    #resp = drvr.update_function_env_vars('ESM-ProcessStarter')
    #print(resp)
    #file_list = drvr.find_files(sub_folder='DataFiles/PivotalPath/input')
    #file_list = drvr.find_files(sub_folder='input')
    #for  key in file_list:
    #    print(" FILE NAME: {} SIZE: {}".format(key,drvr.get_object_size(key)))
    #print(file_list)
    #XML_DATA=drvr.read_text_file('PivotalPath_Combined_Output_09-26-2023.xml', load_from= 'xml')
    #print(XML_DATA)
    os._exit(0)
    #drvr.generate_xml()
    #drvr.consolidate_xml()
    #drvr.archive_files(src_dir = 'temp' , target_dir = 'archive')
    #files = drvr.find_files(sub_folder='processed',file_pattern='xml', as_of=dt.datetime(2023,3,27))
    #files = drvr.find_files(sub_folder='DataFiles/OPSAutomation/input')
    #files = drvr.find_files(sub_folder='processed',file_pattern='014010')
    #for f in files:
    #    drvr.delete_file(f,rm_from='error')
    #print(files)

    #files = drvr.find_files(sub_folder='output_consolidated')
    #print("OUTPUT CONSOLIDATED:",files)
    #files = drvr.find_files(sub_folder='input')
    #print("INPUT:",files)
    #drvr.archive_files(src_dir = 'output_consolidated', target_dir = 'processed')
    #drvr.archive_files(src_dir = 'temp', target_dir = 'processed')

    #df=drvr.read_text_file('Portfolio Transparency Report Jan 2023.csv',load_from='input',
    #                       write_to_df=True,header_row=20,from_stream=True, error_bad_lines='warn')
    
    #print (content)
    #df = pd.read_csv(io.StringIO(content),on_bad_lines=False,skiprows=20)
    #print (list(df.columns))
    #print (df)

    #drvr.trigger_lambda(function_name="ESM-OPSAutomationOutputConsolidator",invocation_type='Event', lambda_payload='{"archive":"1","import":"0","source_folder":"all","target_folder":"processed"}')
    #drvr.trigger_lambda(function_name="ESM-OPSAutomationOutputConsolidator",invocation_type='Event', lambda_payload='{"import":"1"}') #cry_files = drvr.find_files(file_pattern='Crystallization',sub_folder = 'processed')
    #print(cry_files)
    #portfolio_db_df = drvr.download_file(object_name='Em Front Portfolio Valuation 2022.xlsx',
    #                                        load_from="input",data_type='xlsx',write_to_df=True,read_tab=None)
    #tabs = list(portfolio_db_df.keys())
    #print(tabs)
    #jan_df =portfolio_db_df[tabs[0]] 
    #print(jan_df)
    #print (list(portfolio_db_df.keys()))
    #portfolio_db_df = portfolio_db_df.loc[portfolio_db_df['Active'] == True]
    #print(portfolio_db_df['Portfolio_Long_Name'])
    #for index,row in portfolio_db_df.iterrows():
        #p_cry_files = [x for x in cry_files if row['Portfolio_Long_Name'] in x]
        #for cry_file in p_cry_files:
        #print(row['Portfolio_Long_Name'], '==> ',[x for x in cry_files if row['Portfolio_Long_Name'] in x])
    #keyword = "InvestcorpTages Vidrio File 2-8-23"
    #drvr.copy_file(src_file=f'{keyword}.xlsx',dest_file=f'{keyword}.xlsx',load_from=None,load_to='archive',rm_src=True)
    #content = drvr.download_file(object_name='Crystallization Template.xlsx',load_from='templates',data_type='bin',read_tab = None,write_to_df=False) 
    #df=pd.read_excel(content)
    #print ("DF:",df)
    #content.seek(0)
    #cry_template=openpyxl.load_workbook(content)
    #print("getFileContent WB ===> ",cry_template)

    #cry_template = self.fs.get_file_content(crystal_file,aws_folder='templates')
    #portfolio_positions = cry_template["Portfolio Positions"]
    #portfolio_accounts = cry_template["Portfolio Accounts"]
    #portfolio_crystallization = cry_template["Portfolio Crystallization"]

    #cry_template.save("C:/Temp/CryTemplate.xlsx")

    #cry_file_name = '20230214:15:30:32_K2 2023-02-06 Franklin K2 Athena Risk Premia UCITS Fund Crystallization 2023-02-14-15-14-07.xlsx'
    #content = drvr.download_file(object_name=cry_file_name,load_from='output',data_type='bin',read_tab = None,write_to_df=False) 
    #df=pd.read_excel(content)
    #print ("DF:",df)
    #content.seek(0)
    #cry_file=openpyxl.load_workbook(content)
    #print("getFileContent WB ===> ",cry_file)

    #cry_template = self.fs.get_file_content(crystal_file,aws_folder='templates')
    #portfolio_positions = cry_file["Portfolio Positions"]
    #portfolio_accounts = cry_file["Portfolio Accounts"]
    #portfolio_crystallization = cry_file["Portfolio Crystallization"]

    #cry_file.save("C:/Temp/CryFile.xlsx")

    #print(portfolio_positions)

    #df = pd.read_excel("C:/Temp/Crystallization Template.xlsx", sheet_name=None)


    #cry_file = drvr.download_file(cry_file_name ,load_from='output',write_to_df=True,read_tab=None,type='xlsx',data_type="bin")
    #portf_cry = cry_file["Portfolio Crystallization"]
    #pos_cry = cry_file["Portfolio Positions"]
    #acct_cry = cry_file["Portfolio Accounts"]
    #print ("PORTFOLIO CRY : " , portf_cry)
    #print ("PORTFOLIO POSITIONS: ", pos_cry)
    #print ("PORTFOLIO ACCT: ", acct_cry)

    #drvr = S3_Client(source='OPSAUTOMATION')
    #res = drvr.create_subfolder('PivotalPath/output/Test')
    #res = drvr.create_subfolder('PivotalPath/input')
    #drvr.upload_file(file_name='G:/Shared drives/Common/Anatoliy/ESM/esm_release_log.xlsx', load_to='output')
    #drvr.upload_file(file_name='/usr/local/share/vidrio/FileDownloads/PivotalPath/InvestcorpTages Vidrio File 2-25-22.xlsx')
    #drvr.upload_file(file_name='/usr/local/share/vidrio/FileDownloads/PivotalPath/output/PivotalPath_Output_Final_01-25-2022.xlsx',load_to='output')
    #drvr.upload_file(file_name='/usr/local/share/vidrio/FileDownloads/PivotalPath/output/PivotalPath_Output_Merged_01-25-2022.xlsx',load_to='output')
    #drvr.upload_file(file_name='/usr/local/share/vidrio/FileDownloads/PivotalPath/output/PivotalPath_Output_01-25-2022.xlsx',load_to='output')

    #print("COPY FILE:")
    #status = drvr.copy_file(src_file='20220926_2949_Portfolio_Investments.xml',dest_file='20220926_2949_Portfolio_Investments.xml',load_from='input',load_to='archive',rm_src=True)
    #print(status)

    #print("EXCEL VALUATION FILE:")
    #excel_df = drvr.download_file(object_name='Mercer UCITS ALT Valuation Report 20230306.xls',
    # load_from='input',data_type='bin',type=None,
    # write_to_df=True,write_to_file=False,download_to="C:/Temp",skiprows=2,nrows=10)
    # skiprows=2,nrows=5)
    #excel.seek(0)
    #wb=xlrd.open_workbook(file_contents=excel.read())
    #data_sheet = wb.sheet_by_index(0)
    #print(excel_df)
    #wb=openpyxl.load_workbook(excel)
    #wb=openpyxl.load_workbook("G:\Shared drives\PV DEV - WORK ENV\Projects 2022\Tickets\T08400\T08481 - Ops Automation\Test Files\Mercer UCITS ALT Open Trades 20230104.xls")
    #wb=xlrd.open_workbook("G:\Shared drives\PV DEV - WORK ENV\Projects 2022\Tickets\T08400\T08481 - Ops Automation\Test Files\Mercer UCITS ALT Open Trades 20230104.xls")
    #print(wb)
    #ext=os.path.splitext("DataFiles/OPSAutomation/input/Mercer UCITS ALT Open Trades 20230104.xls")[1]
    #print (ext)

    #print("XML POSITIONS FILE:")
    #xml_df = drvr.download_file(object_name='20220523_2948_Portfolio_Investments.xml',load_from='input',data_type='ascii',write_to_df=True,type='xml')
    #print(xml_df['CurrentPeriodClosingMarketValue'])
    #cols=['id','name','investment','manager','peergroup','aumbn']
    #drvr.df_to_xml(df,'InvestcorpTages Vidrio File 2-14-22.xml',root_name="FUNDS",row_name="FUND",columns=cols,index=False,load_to='output')
    #print(df)
    # cloud_path= 's3://com.vidrio.d001-dev/WebDocumentStorage/1300094120227120917272325.enc'
    # file_name='K2 NAV Summary 20220926.xls'
    # drvr.get_file_from_cloud_path(cloud_path,file_name,load_to='input')
    # drvr.reset_AWS_attributes()
    # files=drvr.find_files(sub_folder=drvr.aws_bucket_path+'/input')
    # print ("INPUT FILES: ",files)
        #drvr.write_to_ascii_file(content, '20220926_2950_Portfolio_Investments.xml',load_to='xml')
    #df = drvr.get_current_df()

    #files = os.path.basename(drvr.find_files(file_pattern='InvestcorpTages Vidrio File',sub_folder='DataFiles/PivotalPath/')[0])
    #files = drvr.find_files(sub_folder='DataFiles/OPSAutomation/output')
    #print(files)

    #df=drvr.read_text_file('Custom Portfolio Valuation #113.csv',load_from='input',write_to_df=True,header_row=2)
    #print(df)
     
    #df = drvr.readPasswordProtectedXLFileFromS3('K2Advisors_INDICATIVE NAV_08152022.xlsx', 'K2CADV',load_from='input', read_tab='unspecified')
    #print('DF after readPasswordProtectedXLFileFromS3:',df)

    #password_df = drvr.download_file(object_name="K2 Electron Liquid Global Fund Mapping.xlsx",load_from='mapping',data_type='xlsx',write_to_df=True,read_tab='Password',type='bin') 
    #password = str(password_df["Password"].iloc[0])
    #print(password)
    #df = drvr.readPasswordProtectedXLFileFromS3('K2Advisors_INDICATIVE NAV_12062022.xlsx', 'K2CADV',load_from='input', read_tab=None)
    #non_relevant_tabs = ["FX rate", "Master Data", "PSAM", "Pyxis", "TCORP"]
    #equity_tabs = [tab for tab in df.keys() if tab not in non_relevant_tabs]
    #for tab in equity_tabs:
    #    print('DF after readPasswordProtectedXLFileFromS3: ',tab,'\n',df[tab])
    #drvr.copy_file_direct('DataFiles/OPSAutomation/input/Mercer DAS Open Trades 20221229.xlsx',dest_file='Mercer DAS Open Trades 20221229.xlsx',load_to='archive',rm_src=True)