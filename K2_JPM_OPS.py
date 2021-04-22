#!python3
# JPM Franklin Daily - Crystallization Script

import os
import datetime
import numpy as np
import pandas as pd
import openpyxl # version 2.5.14
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border, NamedStyle, numbers
from datetime import timedelta
import shutil
import sys
import xml.etree.ElementTree as ET
import csv
import xlrd
from Mail import Mail
#from Utils import *

### code for converting xls to XLSX
def xls_to_xlsx(src_file_path, dst_file_path):
    book_xls = xlrd.open_workbook(src_file_path)
    book_xlsx = Workbook()

    sheet_names = book_xls.sheet_names()
    for sheet_index, sheet_name in enumerate(sheet_names):
        sheet_xls = book_xls.sheet_by_name(sheet_name)
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_name
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)

        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row = row+1 , column = col+1).value = sheet_xls.cell_value(row, col)

    book_xlsx.save(dst_file_path)

def K2_JPM_OPS(working_date,folder_path,os,ucits_portfolio_positions_id,franklin_portfolio_positions_id):
    # gather all working XML files in a list
    if (os=='Windows'):
        xml_files = r'G:/Shared drives/Operations/K2/JPM - Daily/FTP Files/'
    else:
        xml_files = folder_path
    today = datetime.date.today()
    todayStr = today.strftime('%Y-%m-%d')
    #working_date = input('Please enter the date of working file in YYYYMMDD format. e.g. 20200729 for July 29, 2020.\n')
    working_date_datetime = datetime.datetime.strptime(working_date, '%Y%m%d')
    working_date_formatted = working_date_datetime.strftime('%Y-%m-%d')
    #ucits_portfolio_positions_id = [2948, 2949, 2950, 4489, 5689, 99129]
    #franklin_portfolio_positions_id = [99045, 99064]
    relevant_files = [file for file in os.listdir(xml_files) if working_date in file and ".xml" in file]

    ucits_portfolio_position_files = []
    for file in relevant_files:
        for id in ucits_portfolio_positions_id:
            if str(id) in file:
                    ucits_portfolio_position_files.append(file)
            else:
                pass

    franklin_portfolio_position_files = []
    for file in relevant_files:
        for id in franklin_portfolio_positions_id:
            if str(id) in file:
                franklin_portfolio_position_files.append(file)
            else:
                pass

    ucits_position_position_files_long = []
    franklin_position_position_files_long = []

    for filename in ucits_portfolio_position_files:
        if (os=='Windows'):
            ucits_position_position_files_long.append('G:/Shared drives/Operations/K2/JPM - Daily/FTP Files/' + filename)
        else:
            ucits_position_position_files_long.append(f'{folder_path}/' + filename)

    for filename in franklin_portfolio_position_files:
        if (os=='Windows'):
            franklin_position_position_files_long.append('G:/Shared drives/Operations/K2/JPM - Daily/FTP Files/' + filename)
        else:
            franklin_position_position_files_long.append(f'{folder_path}/' + filename)

    # go through FTAF NAV and convert data
    if (os=='Windows'):
        ftaf_nav_dir = r'G:/Shared drives/Operations/K2/JPM - Daily/Equity files/'
    else:
        ftaf_nav_dir = folder_path

    ftaf_nav_file_list = [file for file in os.listdir(ftaf_nav_dir) if (working_date in file and 'FTAF' in file)]
    if (os=='Windows'):
        ftaf_nav_file_long = r'G:/Shared drives/Operations/K2/JPM - Daily/Equity files/' + ftaf_nav_file_list[0]
    else:
        ftaf_nav_file_long = f'{folder_path}/' + ftaf_nav_file_list[0]
    print ("K2_JPM_OPS ==> ftaf_nav_file_long : ", ftaf_nav_file_long)
    ftaf_nav_file_long_xlsx = ftaf_nav_file_long[:-4] + '.xlsx'

    xls_to_xlsx(ftaf_nav_file_long, ftaf_nav_file_long_xlsx)

    ftaf_nav = openpyxl.load_workbook(ftaf_nav_file_long_xlsx)
    ftaf_sheet = ftaf_nav[ftaf_nav_file_list[0][:-4]]

    def stringornot(cellval):
        if type(cellval) == str:
            cellval = float(cellval.replace(',',''))
            return cellval
        else:
            return cellval

    for row in range(2, ftaf_sheet.max_row + 1):
        ftaf_sheet['A{}'.format(row)].number_format = numbers.FORMAT_DATE_XLSX14
        ftaf_sheet['I{}'.format(row)].value = float(ftaf_sheet['I{}'.format(row)].value)
        ftaf_sheet['K{}'.format(row)].value = float(ftaf_sheet['K{}'.format(row)].value)
        ftaf_sheet['L{}'.format(row)].value = float(ftaf_sheet['L{}'.format(row)].value)
        ftaf_sheet['M{}'.format(row)].value = numbers.FORMAT_GENERAL
        ftaf_sheet['O{}'.format(row)].number_format = numbers.FORMAT_DATE_XLSX14
        ftaf_sheet['P{}'.format(row)].value = stringornot(ftaf_sheet['P{}'.format(row)].value)
        ftaf_sheet['Q{}'.format(row)].value = stringornot(ftaf_sheet['Q{}'.format(row)].value)
        ftaf_sheet['R{}'.format(row)].value = stringornot(ftaf_sheet['R{}'.format(row)].value)
        ftaf_sheet['S{}'.format(row)].value = stringornot(ftaf_sheet['S{}'.format(row)].value)
        ftaf_sheet['T{}'.format(row)].value = stringornot(ftaf_sheet['T{}'.format(row)].value)
        ftaf_sheet['U{}'.format(row)].value = stringornot(ftaf_sheet['U{}'.format(row)].value)
        ftaf_sheet['V{}'.format(row)].value = stringornot(ftaf_sheet['V{}'.format(row)].value)
        ftaf_sheet['W{}'.format(row)].value = stringornot(ftaf_sheet['W{}'.format(row)].value)
        ftaf_sheet['X{}'.format(row)].value = stringornot(ftaf_sheet['X{}'.format(row)].value)

    # go through K2 NAV - data is already converted
    if (os=='Windows'):
        k2_nav_dir = r'G:/Shared drives/Operations/K2/JPM - Daily/Equity files/'
    else:
        k2_nav_dir = folder_path
    k2_nav_file_list = [file for file in os.listdir(k2_nav_dir) if (working_date in file and 'K2' in file)]
    if (os=='Windows'):
        k2_nav_file_long = r'G:/Shared drives/Operations/K2/JPM - Daily/Equity files/' + k2_nav_file_list[0]
    else:
        k2_nav_file_long = f'{folder_path}/' + k2_nav_file_list[0]
    k2_nav_file_long_xlsx = k2_nav_file_long[:-4] + '.xlsx'

    xls_to_xlsx(k2_nav_file_long, k2_nav_file_long_xlsx)

    k2_nav = openpyxl.load_workbook(k2_nav_file_long_xlsx)
    print ("K2_JPM_OPS ==> k2_nav_file_list[0] : ", k2_nav_file_list[0],k2_nav_file_list[0][:-4])
    k2_sheet = k2_nav[k2_nav_file_list[0][:-4]]

    def stringornot(cellval):
        if type(cellval) == str:
            cellval = float(cellval.replace(',',''))
            return cellval
        else:
            return cellval

    for row in range(2, k2_sheet.max_row + 1):
        k2_sheet['A{}'.format(row)].number_format = numbers.FORMAT_DATE_XLSX14
        k2_sheet['F{}'.format(row)].value = stringornot(k2_sheet['F{}'.format(row)].value)
        k2_sheet['H{}'.format(row)].value = stringornot(k2_sheet['H{}'.format(row)].value)
        k2_sheet['I{}'.format(row)].value = stringornot(k2_sheet['I{}'.format(row)].value)
        k2_sheet['L{}'.format(row)].value = stringornot(k2_sheet['L{}'.format(row)].value)
        k2_sheet['M{}'.format(row)].value = stringornot(k2_sheet['M{}'.format(row)].value)
        k2_sheet['N{}'.format(row)].value = stringornot(k2_sheet['N{}'.format(row)].value)
        k2_sheet['O{}'.format(row)].value = stringornot(k2_sheet['O{}'.format(row)].value)
        k2_sheet['P{}'.format(row)].value = stringornot(k2_sheet['P{}'.format(row)].value)

    ########################################################################################################################################
    # December 2, 2020 - Changed compiled sheet legacy code to updated

    # parse through XML files and compile into a master spreadsheet
    position_files_long = []

    xml_investments = {}
    missing_xml_investments = {}

    for filename in relevant_files:
        if (os=='Windows'):
            position_files_long.append('G:/Shared drives/Operations/K2/JPM - Daily/FTP Files/' + filename)
        else:
            position_files_long.append(f'{folder_path}/' + filename)

    for file in position_files_long:
        print ("Positions xml File: ",file)
        tree = ET.parse(file)
        root = tree.getroot() # root -> K2PortfolioInvestment Reporting
        for subnode in root: # subnode -> Portfolio
            for subnode2 in subnode.findall('Investment'): # subnode2 -> Investment
                pull_data = []
                pull_data.append(subnode.find('Period').text)
                pull_data.append(subnode2.find('InvestmentName').text)
                for subnode3 in subnode2.findall('BaseCCY'):
                    pull_data.append(subnode3.find('CurrentPeriodSubscriptions').text)
                    pull_data.append(subnode3.find('CurrentPeriodRedemptions').text)
                    pull_data.append(subnode3.find('CurrentPeriodPnL').text)
                    pull_data.append(subnode3.find('CurrentPeriodClosingMarketValue').text)
                xml_investments.setdefault(subnode2.find('InvestmentID').text, pull_data)

    compiled_file_template = Workbook()
    if (os=='Windows'):
        compiled_file_copy = r'C:/Users/Walter/Desktop/Vidrio/' + working_date_formatted + ' Compiled File Template ' + todayStr + '.xlsx'
    else:
        compiled_file_copy = f'{folder_path}/templates/' + working_date_formatted + ' Compiled File Template ' + todayStr + '.xlsx'
    compiled_file_sheet = compiled_file_template.create_sheet('Sheet1', 0)
    compiled_file_sheet['A1'].value = 'Period'
    compiled_file_sheet['B1'].value = 'InvestmentID'
    compiled_file_sheet['C1'].value = 'InvestmentName'
    compiled_file_sheet['D1'].value = 'CurrentPeriodSubscriptions'
    compiled_file_sheet['E1'].value = 'CurrentPeriodRedemptions'
    compiled_file_sheet['F1'].value = 'CurrentPeriodPnL'
    compiled_file_sheet['G1'].value = 'CurrentPeriodClosingMarketValue'

    compiled_row_counter = 2 # row in excel
    for index, keys in enumerate(xml_investments):
        compiled_file_sheet['A{}'.format(compiled_row_counter)].value = xml_investments[keys][0] # period
        compiled_file_sheet['A{}'.format(compiled_row_counter)].number_format = numbers.FORMAT_DATE_XLSX14
        compiled_file_sheet['B{}'.format(compiled_row_counter)].value = float(keys) # investment id
        compiled_file_sheet['C{}'.format(compiled_row_counter)].value = xml_investments[keys][1] # investment name
        compiled_file_sheet['D{}'.format(compiled_row_counter)].value = float(xml_investments[keys][2]) # currentPeriodSub
        compiled_file_sheet['E{}'.format(compiled_row_counter)].value = float(xml_investments[keys][3]) # currentPeriodRedemp
        compiled_file_sheet['F{}'.format(compiled_row_counter)].value = float(xml_investments[keys][4]) # currentPeriodPnL
        compiled_file_sheet['G{}'.format(compiled_row_counter)].value = float(xml_investments[keys][5]) # currentPeriodClosingMarket
        compiled_row_counter += 1

    compiled_file_template.save(compiled_file_copy)
    ##################################################################################################################################################

    #########################################
    # create exception report
    if (os=='Windows'):
        missing_account_directory = r'G:/Shared drives/Operations/K2/JPM - Daily/'
    else:
        missing_account_directory = f'{folder_path}/processing/'

    missing_account = Workbook()
    missing_account_sheet = missing_account.create_sheet('Missing Accounts', 0)
    missing_account_sheet['A1'].value = 'InvestmentID'
    missing_account_sheet['B1'].value = 'Investment Name'
    missing_account_sheet['C1'].value = 'Portfolio'

    ### missing accounts for franklin
    if (os=='Windows'):
        franklin_mapping_file_ext = r'G:/Shared drives/Operations/K2/JPM - Daily/Processing/JPM Franklin/JPM Franklin Mapping.xlsx'
        franklin_mapping_backup = r'G:/Shared drives/Operations/K2/JPM - Daily/Processing/JPM Franklin/' + 'JPM Franklin Mapping Backup pre ' + working_date_formatted + '.xlsx'
    else:
        franklin_mapping_file_ext = f'{folder_path}/mapping/JPM Franklin Mapping.xlsx'
        franklin_mapping_backup = f'{folder_path}/backup/' + 'JPM Franklin Mapping Backup pre ' + working_date_formatted + '.xlsx'
    franklin_mapping_file = openpyxl.load_workbook(franklin_mapping_file_ext)
    franklin_mapping_sheet = franklin_mapping_file['JPM Franklin']
    franklin_mapping_notes = franklin_mapping_file['Notes']
    franklin_update_counter = 0

    franklin_account_list = []

    for row in range(2, franklin_mapping_sheet.max_row):
        if franklin_mapping_sheet['A{}'.format(row)].value and franklin_mapping_sheet['E{}'.format(row)].value == True:
            franklin_account_list.append(str(franklin_mapping_sheet['A{}'.format(row)].value))

    filler = 2
    missing_counter = 0
    missing_status=""
    for row in range(2, k2_sheet.max_row+1):
        id_long = (str(k2_sheet['C{}'.format(row)].value) + str(k2_sheet['D{}'.format(row)].value))
        if id_long not in franklin_account_list:
            print(id_long, k2_sheet['B{}'.format(row)].value, 'not in Franklin mapping file.')
            missing_status = missing_status + id_long + k2_sheet['B{}'.format(row)].value+'\n'
            missing_account_sheet['A{}'.format(filler)].value = id_long
            missing_account_sheet['B{}'.format(filler)].value = k2_sheet['B{}'.format(row)].value
            missing_account_sheet['C{}'.format(filler)].value = 'JPM Franklin'
            filler += 1
            missing_counter += 1

    if missing_counter > 0 and len(missing_status) > 0:
        #send an email here...
        m=Mail("Mail",'anatoliy.shor@vidrio.com','anatoliy.shor@vidrio.com','gygmy.gonnot@vidrio.com','K2 OPS JPM Error -- Accounts Missing from Franklin file')
        m.set_body(missing_status)
        m.send_text_email()


    if missing_counter == 0:
        print('All account numbers were found in the Franklin mapping file.\n')
    else:
        missing_account.save(missing_account_directory + 'JPM Franklin Missing Accounts ' + working_date_formatted + '.xlsx')

    ### missing acounts for ucits
    if (os=='Windows'):
        ucits_mapping_file_ext = r'G:/Shared drives/Operations/K2/JPM - Daily/Processing/JPM Ucits/JPM UCITS Mapping.xlsx'
        ucits_mapping_backup = r'G:/Shared drives/Operations/K2/JPM - Daily/Processing/JPM Ucits/' + 'JPM UCITS Mapping Backup pre ' + working_date_formatted + '.xlsx'
    else:
        ucits_mapping_file_ext = f'{folder_path}/mapping/JPM UCITS Mapping.xlsx'
        ucits_mapping_backup = f'{folder_path}/backup/JPM UCITS Mapping Backup pre ' + working_date_formatted + '.xlsx'
    ucits_mapping_file = openpyxl.load_workbook(ucits_mapping_file_ext)
    ucits_mapping_sheet = ucits_mapping_file['JPM Daily']
    ucits_mapping_notes = ucits_mapping_file['Notes']
    ucits_update_counter = 0

    ucits_account_list = []

    for row in range(2, ucits_mapping_sheet.max_row):
        if ucits_mapping_sheet['A{}'.format(row)].value and ucits_mapping_sheet['E{}'.format(row)].value == True:
            ucits_account_list.append(str(ucits_mapping_sheet['A{}'.format(row)].value))

    missing_status=""
    for row in range(2, ftaf_sheet.max_row+1):
        id_long = (str(ftaf_sheet['F{}'.format(row)].value) + str(ftaf_sheet['C{}'.format(row)].value))
        if id_long not in ucits_account_list:
            print(id_long, ftaf_sheet['E{}'.format(row)].value, 'not in UCITS mapping file.')
            missing_status = missing_status + id_long + "  " + ftaf_sheet['E{}'.format(row)].value+'\n'
            missing_account_sheet['A{}'.format(filler)].value = id_long
            missing_account_sheet['B{}'.format(filler)].value = ftaf_sheet['E{}'.format(row)].value
            missing_account_sheet['C{}'.format(filler)].value = 'JPM UCITS'
            filler += 1
            missing_counter += 1

    if missing_counter > 0 and len(missing_status) > 0:
        #send an email here...
        print (missing_status)
        m=Mail("Mail",'anatoliy.shor@vidrio.com','anatoliy.shor@vidrio.com','gygmy.gonnot@vidrio.com','K2 OPS JPM Error -- Accounts Missing from UCITS file')
        m.set_body(missing_status)
        m.send_text_email()

    if missing_counter == 0:
        print('All account numbers were found in both JPM Franklin and UCITS mapping file.\n')
    else:
        missing_account.save(missing_account_directory + 'JPM Missing Port Accounts ' + working_date_formatted + '.xlsx')
        sys.exit() # stop for exception report

    ## missing accounts for XML Investments
    xml_investments_list = []

    # add portfolio posiions to xml list
    for row in range(2, franklin_mapping_sheet.max_row):
        if franklin_mapping_sheet['A{}'.format(row)].value and franklin_mapping_sheet['E{}'.format(row)].value == False:
            xml_investments_list.append(str(franklin_mapping_sheet['A{}'.format(row)].value))

    for row in range(2, ucits_mapping_sheet.max_row):
        if ucits_mapping_sheet['A{}'.format(row)].value and ucits_mapping_sheet['E{}'.format(row)].value == False:
            xml_investments_list.append(str(ucits_mapping_sheet['A{}'.format(row)].value))

    # clear missing account for xml investments
    missing_account_sheet.delete_rows(2, 1000)

    filler = 2
    missing_counter = 0

    for ids in xml_investments:
        if ids not in xml_investments_list and float(xml_investments[ids][5]) > 0:
            print(ids, xml_investments[ids][1], 'not in either mapping file.')
            missing_account_sheet['A{}'.format(filler)].value = ids
            missing_account_sheet['B{}'.format(filler)].value = xml_investments[ids][1]
            filler += 1
            missing_counter += 1

    if missing_counter == 0:
        print('All investment account numbers were found in both JPM Franklin and UCITS mapping file.\n')
    else:
        missing_account.save(missing_account_directory + 'JPM Missing Investment Accounts ' + working_date_formatted + '.xlsx')
        sys.exit() # stop for exception report


    #########################
    # have to work on both compiled file and k2 file now to build crystal, trans, acct perf ->
    # work on acct perf first

    print("The last crystallization date ran was " + franklin_mapping_notes['B2'].value.strftime('%m-%d-%y') + ".")
    continue_run = True
    continue_run = True

    #run_decision = input("Would you like to continue? (y/n)\n")
    #while continue_run:
    #    if run_decision == 'y':
    #       continue_run = False
    #    elif run_decision == 'n':
    #        sys.exit()

    shutil.copyfile(franklin_mapping_file_ext, franklin_mapping_backup)
    shutil.copyfile(ucits_mapping_file_ext, ucits_mapping_backup)

    franklin_mapping_df = pd.read_excel(franklin_mapping_file_ext)
    ucits_mapping_df = pd.read_excel(ucits_mapping_file_ext)

    franklin_portfolios = pd.Series(franklin_mapping_df.Portfolio_Column.unique())
    ucits_portfolios = pd.Series(ucits_mapping_df.Portfolio_Column.unique())



    #days = [1, 2, 3]
#    days = [0]
#    if working_date_datetime.day in days and (working_date_datetime.weekday() < 6):
#        print("Update needed on shares as it is the first of the month or after.\n")
#        bardin_hill_arb = float(input("Please enter the shares for Franklin K2 Bardin Hill Arbitrage UCITS Fund.\n"))
#        ucits_mapping_sheet['H{}'.format(ucits_mapping_sheet.max_row - 4)].value = bardin_hill_arb
#        chilton_equity_ls = float(input("Please enter the shares for Franklin K2 Chilton Equity Long Short UCITS Fund.\n"))
#        ucits_mapping_sheet['H{}'.format(ucits_mapping_sheet.max_row - 3)].value = chilton_equity_ls
#        electron_global = float(input("Please enter the shares for Franklin K2 Electron Global UCITS Fund.\n"))
#        ucits_mapping_sheet['H{}'.format(ucits_mapping_sheet.max_row - 2)].value = electron_global
#        ellington_structured_credit = float(input("Please enter the shares for Franklin K2 Ellington Structured Credit UCITS Fund.\n"))
#        ucits_mapping_sheet['H{}'.format(ucits_mapping_sheet.max_row - 1)].value = ellington_structured_credit
#        wellington_tech_ls = float(input("Please enter the shares for Franklin K2 Wellington Technology Long Short UCITS Fund.\n"))
#        ucits_mapping_sheet['H{}'.format(ucits_mapping_sheet.max_row)].value = wellington_tech_ls
#        alt_strat_pool = float(input("Please enter the shares for Franklin K2 Alternative Strategies Fund.\n"))
#        franklin_mapping_sheet['H{}'.format(franklin_mapping_sheet.max_row - 1)].value = alt_strat_pool
#        ls_pool = float(input("Please enter the shares for Franklin K2  Long Short Fund.\n"))
#        franklin_mapping_sheet['H{}'.format(franklin_mapping_sheet.max_row)].value = ls_pool

    #--------------------------------------------------------------------------------
    # crystallization

    if (os=='Windows'):
        crystal_file = r'G:/Shared drives/Operations/K2/JPM - Daily/Processing/Crystallization Template.xlsx'
    else:
        crystal_file = f'{folder_path}/templates/Crystallization Template.xlsx'
    crystal = openpyxl.load_workbook(crystal_file)

    #--- portfolio crystal sheet first
    portfolio_crystal = crystal['Portfolio Crystallization']
    portfolio_crystal.delete_rows(2, 1000)

    idx = (working_date_datetime.weekday() + 1) % 7
    sat = working_date_datetime - datetime.timedelta(7+idx-6)

    cry_start_date = franklin_mapping_notes['B2'].value
    cry_start_date += timedelta(days=1)

    portfolio_crystal_counter = 2
    #portfolio_funds = []

    #for row in range(2, ucits_mapping_sheet.max_row + 1):
    #    if ucits_mapping_sheet['D{}'.format(row)].value and ucits_mapping_sheet['D{}'.format(row)].value not in portfolio_funds:
    #        portfolio_funds.append(ucits_mapping_sheet['D{}'.format(row)].value)

    #for index, value in enumerate(portfolio_funds):
    for index in ucits_portfolios:
        portfolio_crystal['A{}'.format(portfolio_crystal_counter)].value = index + ' - ' + working_date_formatted # crystallization reference
        portfolio_crystal['C{}'.format(portfolio_crystal_counter)].value = index # Portfolio Long Name
        #if working_date_datetime.weekday() == 0:
        #    portfolio_crystal['D{}'.format(portfolio_crystal_counter)].value = sat # Starting Date
        #    portfolio_crystal['D{}'.format(portfolio_crystal_counter)].number_format = numbers.FORMAT_DATE_XLSX14
        #else:
        #    portfolio_crystal['D{}'.format(portfolio_crystal_counter)].value = working_date_datetime # starting date
        #    portfolio_crystal['D{}'.format(portfolio_crystal_counter)].number_format = numbers.FORMAT_DATE_XLSX14
        portfolio_crystal['D{}'.format(portfolio_crystal_counter)].value = cry_start_date # Starting Date
        portfolio_crystal['D{}'.format(portfolio_crystal_counter)].number_format = numbers.FORMAT_DATE_XLSX14
        portfolio_crystal['E{}'.format(portfolio_crystal_counter)].value = working_date_datetime # Ending Date
        portfolio_crystal['E{}'.format(portfolio_crystal_counter)].number_format = numbers.FORMAT_DATE_XLSX14
        portfolio_crystal['F{}'.format(portfolio_crystal_counter)].value = working_date_datetime # Statement Date
        portfolio_crystal['F{}'.format(portfolio_crystal_counter)].number_format = numbers.FORMAT_DATE_XLSX14
        portfolio_crystal['G{}'.format(portfolio_crystal_counter)].value = 'Final' # Status
        portfolio_crystal_counter += 1

    ## portfolio crystal done
    portfolio_positions = crystal['Portfolio Positions']
    portfolio_positions.delete_rows(2, 1000)

    portfolio_accounts = crystal['Portfolio Accounts']
    portfolio_accounts.delete_rows(2, 1000)

    portfolio_positions_counter = 2
    portfolio_accounts_counter = 2

    # bardin_fund = 0
    # chilton_fund = 0
    # electron_fund = 0
    # ellington_fund = 0
    # wellington_fund = 0
    # emso_fund = 0

    for row in range(2, compiled_file_sheet.max_row + 1):
        for mapping_row in range(2, ucits_mapping_sheet.max_row):
            if ucits_mapping_sheet['E{}'.format(mapping_row)].value == False and ucits_mapping_sheet['M{}'.format(mapping_row)].value == True: # for port positions boolean
                if compiled_file_sheet['B{}'.format(row)].value == ucits_mapping_sheet['A{}'.format(mapping_row)].value:
                    trade_value = compiled_file_sheet['D{}'.format(row)].value + compiled_file_sheet['E{}'.format(row)].value
                    yesterday_nav = ucits_mapping_sheet['G{}'.format(mapping_row)].value
                    trade_shares = trade_value / yesterday_nav
                    currentPeriodPnL = compiled_file_sheet['F{}'.format(row)].value
                    today_shares = ucits_mapping_sheet['H{}'.format(mapping_row)].value + trade_shares

                    portfolio_positions['A{}'.format(portfolio_positions_counter)].value = ucits_mapping_sheet['D{}'.format(mapping_row)].value + ' - ' + working_date_formatted # crystal reference
                    portfolio_positions['B{}'.format(portfolio_positions_counter)].value = ucits_mapping_sheet['B{}'.format(mapping_row)].value # investor acct
                    portfolio_positions['C{}'.format(portfolio_positions_counter)].value = ucits_mapping_sheet['C{}'.format(mapping_row)].value # ivstment acct
                    portfolio_positions['D{}'.format(portfolio_positions_counter)].value = today_shares # units closing
                    portfolio_positions['F{}'.format(portfolio_positions_counter)].value = compiled_file_sheet['G{}'.format(row)].value # investment closing bal
                    portfolio_positions['E{}'.format(portfolio_positions_counter)].value = portfolio_positions['F{}'.format(portfolio_positions_counter)].value / today_shares # closing nav
                    portfolio_positions['G{}'.format(portfolio_positions_counter)].value = 1 # invstor closing fx rate
                    portfolio_positions['H{}'.format(portfolio_positions_counter)].value = portfolio_positions['F{}'.format(portfolio_positions_counter)].value # invstor closing bal
                    portfolio_positions['I{}'.format(portfolio_positions_counter)].value = 1 # port closing fx
                    portfolio_positions['J{}'.format(portfolio_positions_counter)].value = portfolio_positions['F{}'.format(portfolio_positions_counter)].value # port closing bal

###NEED TO FIX THIS HARDCODED ACCOUNTS
# addressed below in line 547
                    # if '3019415200' in str(compiled_file_sheet['B{}'.format(row)].value): # bardin
                    #     bardin_fund = compiled_file_sheet['F{}'.format(row)].value
                    # elif '3019414500' in str(compiled_file_sheet['B{}'.format(row)].value): # chilton
                    #     chilton_fund = compiled_file_sheet['F{}'.format(row)].value
                    # elif '3019417300' in str(compiled_file_sheet['B{}'.format(row)].value): # electron
                    #     electron_fund = compiled_file_sheet['F{}'.format(row)].value
                    # elif '3019413600' in str(compiled_file_sheet['B{}'.format(row)].value): # ellington
                    #     ellington_fund = compiled_file_sheet['F{}'.format(row)].value
                    # elif '3019416100' in str(compiled_file_sheet['B{}'.format(row)].value): # wellington
                    #     wellington_fund = compiled_file_sheet['F{}'.format(row)].value
                    # elif '4061618300' in str(compiled_file_sheet['B{}'.format(row)].value): # emso
                    #     emso_fund = compiled_file_sheet['F{}'.format(row)].value

                    # rewrite mapping file edit
                    ucits_mapping_sheet['F{}'.format(mapping_row)].value = working_date_datetime # update previous NAV
                    ucits_mapping_sheet['F{}'.format(mapping_row)].number_format = numbers.FORMAT_DATE_XLSX14
                    ucits_mapping_sheet['J{}'.format(mapping_row)].value = float(ucits_mapping_sheet['G{}'.format(mapping_row)].value) # update previous NAV
                    ucits_mapping_sheet['K{}'.format(mapping_row)].value = float(ucits_mapping_sheet['H{}'.format(mapping_row)].value) # update previous O/S
                    ucits_mapping_sheet['L{}'.format(mapping_row)].value = float(ucits_mapping_sheet['I{}'.format(mapping_row)].value) # update previous Total Assets

                    ucits_mapping_sheet['G{}'.format(mapping_row)].value = float(portfolio_positions['E{}'.format(portfolio_positions_counter)].value) # update for today NAV
                    ucits_mapping_sheet['H{}'.format(mapping_row)].value = float(today_shares) # update for today O/S
                    ucits_mapping_sheet['I{}'.format(mapping_row)].value = float(portfolio_positions['F{}'.format(portfolio_positions_counter)].value) # update for today Total Assets

                    portfolio_positions_counter += 1
                    ucits_update_counter += 1

    for mapping_row in range(2, ucits_mapping_sheet.max_row  + 1):
        for ftaf_row in range(2, ftaf_sheet.max_row + 1):
            if ucits_mapping_sheet['E{}'.format(mapping_row)].value == True and ucits_mapping_sheet['M{}'.format(mapping_row)].value == True and ftaf_sheet['E{}'.format(ftaf_row)].value == ucits_mapping_sheet['C{}'.format(mapping_row)].value:
                portfolio_accounts['A{}'.format(portfolio_accounts_counter)].value = ucits_mapping_sheet['D{}'.format(mapping_row)].value + ' - ' + working_date_formatted  # crystal ref
                portfolio_accounts['B{}'.format(portfolio_accounts_counter)].value = ucits_mapping_sheet['C{}'.format(mapping_row)].value  # Port Acct
                portfolio_accounts['C{}'.format(portfolio_accounts_counter)].value = None # opening Units
                try:
                    portfolio_accounts['D{}'.format(portfolio_accounts_counter)].value = ftaf_sheet['R{}'.format(ftaf_row)].value / ftaf_sheet['X{}'.format(ftaf_row)].value # Opening Nav
                except ZeroDivisionError:
                    portfolio_accounts['D{}'.format(portfolio_accounts_counter)].value = ucits_mapping_sheet['G{}'.format(mapping_row)].value # initialized before in mapping file
                portfolio_accounts['E{}'.format(portfolio_accounts_counter)].value = None # Opening FX Rate
                portfolio_accounts['H{}'.format(portfolio_accounts_counter)].value = 0 # FX PnL
                portfolio_accounts['I{}'.format(portfolio_accounts_counter)].value = 0 # FX Hedge PnL
                portfolio_accounts['J{}'.format(portfolio_accounts_counter)].value = 0 # Interest income
                portfolio_accounts['K{}'.format(portfolio_accounts_counter)].value = 0 # int expense
                portfolio_accounts['L{}'.format(portfolio_accounts_counter)].value = 0 # admin fe
                portfolio_accounts['M{}'.format(portfolio_accounts_counter)].value = 0 # custody fee
                portfolio_accounts['N{}'.format(portfolio_accounts_counter)].value = 0 # misc expense
                portfolio_accounts['O{}'.format(portfolio_accounts_counter)].value = 0 # acct expenses
                portfolio_accounts['P{}'.format(portfolio_accounts_counter)].value = 0 # management fee
                portfolio_accounts['Q{}'.format(portfolio_accounts_counter)].value = 0 # perf fee
                portfolio_accounts['R{}'.format(portfolio_accounts_counter)].value = 0 # acct fees
                portfolio_accounts['S{}'.format(portfolio_accounts_counter)].value = ftaf_sheet['P{}'.format(ftaf_row)].value # closing bal
                portfolio_accounts['T{}'.format(portfolio_accounts_counter)].value = ftaf_sheet['W{}'.format(ftaf_row)].value # closing units
                portfolio_accounts['F{}'.format(portfolio_accounts_counter)].value = portfolio_accounts['D{}'.format(portfolio_accounts_counter)].value * portfolio_accounts['T{}'.format(portfolio_accounts_counter)].value #Adj Opening Bal
                portfolio_accounts['G{}'.format(portfolio_accounts_counter)].value = portfolio_accounts['S{}'.format(portfolio_accounts_counter)].value - portfolio_accounts['F{}'.format(portfolio_accounts_counter)].value # Invst PnL
                portfolio_accounts['U{}'.format(portfolio_accounts_counter)].value = ftaf_sheet['I{}'.format(ftaf_row)].value # closing nav
                portfolio_accounts['V{}'.format(portfolio_accounts_counter)].value = ftaf_sheet['P{}'.format(ftaf_row)].value / ftaf_sheet['S{}'.format(ftaf_row)].value

                # rewrite mapping file edit
                ucits_mapping_sheet['F{}'.format(mapping_row)].value = working_date_datetime # update previous NAV
                ucits_mapping_sheet['F{}'.format(mapping_row)].number_format = numbers.FORMAT_DATE_XLSX14
                ucits_mapping_sheet['J{}'.format(mapping_row)].value = float(ucits_mapping_sheet['G{}'.format(mapping_row)].value) # update previous NAV
                ucits_mapping_sheet['K{}'.format(mapping_row)].value = float(ucits_mapping_sheet['H{}'.format(mapping_row)].value) # update previous O/S
                ucits_mapping_sheet['L{}'.format(mapping_row)].value = float(ucits_mapping_sheet['I{}'.format(mapping_row)].value) # update previous Total Assets

                ucits_mapping_sheet['G{}'.format(mapping_row)].value = portfolio_accounts['U{}'.format(portfolio_accounts_counter)].value # update for today NAV
                ucits_mapping_sheet['H{}'.format(mapping_row)].value = ftaf_sheet['W{}'.format(ftaf_row)].value # update for today O/S
                ucits_mapping_sheet['I{}'.format(mapping_row)].value = ftaf_sheet['P{}'.format(ftaf_row)].value # update for today Total Assets

                portfolio_accounts_counter += 1
                ucits_update_counter += 1


    # initialize investment pools with their PnL
    investment_pool_funds = {}

    # create a dictionary correlating Investment Pool Name: Investment Pool PnL
    for row in range(2, portfolio_positions.max_row + 1):
        if portfolio_positions['B{}'.format(row)].value and portfolio_positions['B{}'.format(row)].value not in investment_pool_funds:
            investment_pool_funds.setdefault(str(portfolio_positions['B{}'.format(row)].value), 0)

    # comparing the id's for portfolio positions in compiled file sheet with mapping sheet
    # we will see if the Investor Name for XML positions is equal, thus adding PnL to the dictionary key
    for row in range(2, compiled_file_sheet.max_row + 1):
        for mapping_row in range(2, ucits_mapping_sheet.max_row):
            if compiled_file_sheet['B{}'.format(row)].value == ucits_mapping_sheet['A{}'.format(mapping_row)].value:
                print("Investment pool funds: ",investment_pool_funds)
                investment_pool_funds[str(ucits_mapping_sheet['B{}'.format(mapping_row)].value)] += compiled_file_sheet['F{}'.format(row)].value

###NEED TO FIX THIS HARDCODED ACCOUNTS
# fix is above in 553
    # investment_pool_PnL = [bardin_fund, chilton_fund, electron_fund, ellington_fund, wellington_fund, emso_fund]

    for num, key in enumerate(investment_pool_funds):
        current_row = ucits_mapping_sheet.max_row - len(investment_pool_funds) + 1 + num
        portfolio_accounts['A{}'.format(portfolio_accounts_counter)].value = str(ucits_mapping_sheet['D{}'.format(current_row)].value) + ' - ' + working_date_formatted # crystal ref
        portfolio_accounts['B{}'.format(portfolio_accounts_counter)].value = str(ucits_mapping_sheet['C{}'.format(current_row)].value) # Port Acct
        portfolio_accounts['C{}'.format(portfolio_accounts_counter)].value = None # opening Units
        portfolio_accounts['D{}'.format(portfolio_accounts_counter)].value = float(ucits_mapping_sheet['G{}'.format(current_row)].value) # Opening Nav
        portfolio_accounts['E{}'.format(portfolio_accounts_counter)].value = None # Opening FX Rate
        portfolio_accounts['F{}'.format(portfolio_accounts_counter)].value = ucits_mapping_sheet['G{}'.format(current_row)].value * ucits_mapping_sheet['H{}'.format(current_row)].value # Adj Opening Bal
        portfolio_accounts['G{}'.format(portfolio_accounts_counter)].value = investment_pool_funds[str(ucits_mapping_sheet['C{}'.format(current_row)].value)] # Invst PnL
        portfolio_accounts['H{}'.format(portfolio_accounts_counter)].value = 0 # FX PnL
        portfolio_accounts['I{}'.format(portfolio_accounts_counter)].value = 0 # FX Hedge PnL
        portfolio_accounts['J{}'.format(portfolio_accounts_counter)].value = 0 # Interest income
        portfolio_accounts['K{}'.format(portfolio_accounts_counter)].value = 0 # int expense
        portfolio_accounts['L{}'.format(portfolio_accounts_counter)].value = 0 # admin fee
        portfolio_accounts['M{}'.format(portfolio_accounts_counter)].value = 0 # custody fee
        portfolio_accounts['N{}'.format(portfolio_accounts_counter)].value = 0 # misc expense
        portfolio_accounts['O{}'.format(portfolio_accounts_counter)].value = 0 # acct expenses
        portfolio_accounts['P{}'.format(portfolio_accounts_counter)].value = 0 # management fee
        portfolio_accounts['Q{}'.format(portfolio_accounts_counter)].value = 0 # perf fee
        portfolio_accounts['R{}'.format(portfolio_accounts_counter)].value = 0 # acct fees
        portfolio_accounts['S{}'.format(portfolio_accounts_counter)].value = portfolio_accounts['F{}'.format(portfolio_accounts_counter)].value + portfolio_accounts['G{}'.format(portfolio_accounts_counter)].value # closing bal
        portfolio_accounts['T{}'.format(portfolio_accounts_counter)].value = ucits_mapping_sheet['H{}'.format(current_row)].value # closing units from mapping file
        portfolio_accounts['U{}'.format(portfolio_accounts_counter)].value = portfolio_accounts['S{}'.format(portfolio_accounts_counter)].value / portfolio_accounts['T{}'.format(portfolio_accounts_counter)].value # closing nav
        portfolio_accounts['V{}'.format(portfolio_accounts_counter)].value = 1

        # rewrite for mapping file edit
        ucits_mapping_sheet['F{}'.format(current_row)].value = working_date_datetime # update previous NAV
        ucits_mapping_sheet['F{}'.format(current_row)].number_format = numbers.FORMAT_DATE_XLSX14
        ucits_mapping_sheet['J{}'.format(current_row)].value = float(ucits_mapping_sheet['G{}'.format(current_row)].value) # update previous NAV
        ucits_mapping_sheet['K{}'.format(current_row)].value = float(ucits_mapping_sheet['H{}'.format(current_row)].value) # update previous O/S
        ucits_mapping_sheet['L{}'.format(current_row)].value = float(ucits_mapping_sheet['I{}'.format(current_row)].value) # update previous Total Assets

        ucits_mapping_sheet['G{}'.format(current_row)].value = portfolio_accounts['U{}'.format(portfolio_accounts_counter)].value # update for today NAV
        ucits_mapping_sheet['H{}'.format(current_row)].value = portfolio_accounts['T{}'.format(portfolio_accounts_counter)].value # update for today O/S
        ucits_mapping_sheet['I{}'.format(current_row)].value = portfolio_accounts['S{}'.format(portfolio_accounts_counter)].value # update for today Total Assets

        portfolio_accounts_counter += 1
        ucits_update_counter += 1

    if (os=='Windows'):
        ucits_crystal_dest = r'G:/Shared drives/Operations/K2/JPM - Daily/Import/' + working_date_formatted + ' UCITS Portfolio Crystallization ' + todayStr + '.xlsx'
    else:
        ucits_crystal_dest = f'{folder_path}/output/' + working_date_formatted + ' UCITS Portfolio Crystallization ' + todayStr + '.xlsx'

    crystal.save(ucits_crystal_dest)

    ## franklin crystal
    portfolio_crystal.delete_rows(2, 1000)
    portfolio_crystal_counter = 2

    # address fix below for legacy code line 730
    # portfolio_funds = ['FTIF Franklin K2 Alternative Strategies Fund',
    #                 'FTIF Franklin K2 Long Short Credit Fund']

    #for index, value in enumerate(portfolio_funds):
    for index in franklin_portfolios:
        portfolio_crystal['A{}'.format(portfolio_crystal_counter)].value = index + ' - ' + working_date_formatted # crystallization reference
        portfolio_crystal['C{}'.format(portfolio_crystal_counter)].value = index # Portfolio Long Name
        #if working_date_datetime.weekday() == 0:
        #    portfolio_crystal['D{}'.format(portfolio_crystal_counter)].value = sat # Starting Date
        #    portfolio_crystal['D{}'.format(portfolio_crystal_counter)].number_format = numbers.FORMAT_DATE_XLSX14
        #else:
        #    portfolio_crystal['D{}'.format(portfolio_crystal_counter)].value = working_date_datetime # starting date
        #    portfolio_crystal['D{}'.format(portfolio_crystal_counter)].number_format = numbers.FORMAT_DATE_XLSX14
        portfolio_crystal['D{}'.format(portfolio_crystal_counter)].value = cry_start_date # Starting Date
        portfolio_crystal['D{}'.format(portfolio_crystal_counter)].number_format = numbers.FORMAT_DATE_XLSX14
        portfolio_crystal['E{}'.format(portfolio_crystal_counter)].value = working_date_datetime # Ending Date
        portfolio_crystal['E{}'.format(portfolio_crystal_counter)].number_format = numbers.FORMAT_DATE_XLSX14
        portfolio_crystal['F{}'.format(portfolio_crystal_counter)].value = working_date_datetime # Statement Date
        portfolio_crystal['F{}'.format(portfolio_crystal_counter)].number_format = numbers.FORMAT_DATE_XLSX14
        portfolio_crystal['G{}'.format(portfolio_crystal_counter)].value = 'Final' # Status
        portfolio_crystal_counter += 1

    portfolio_positions.delete_rows(2, 1000)
    portfolio_accounts.delete_rows(2, 1000)

    portfolio_positions_counter = 2
    portfolio_accounts_counter = 2

    # fix on line 730
    # alt_strat_fund = 0
    # ls_fund = 0

    for row in range(2, compiled_file_sheet.max_row + 1):
        for mapping_row in range(2, franklin_mapping_sheet.max_row):
            if franklin_mapping_sheet['E{}'.format(mapping_row)].value == False and franklin_mapping_sheet['M{}'.format(mapping_row)].value == True: # for port positions boolean
                if compiled_file_sheet['B{}'.format(row)].value == franklin_mapping_sheet['A{}'.format(mapping_row)].value and franklin_mapping_sheet['H{}'.format(mapping_row)].value > 0:
                    trade_value = compiled_file_sheet['D{}'.format(row)].value + compiled_file_sheet['E{}'.format(row)].value
                    yesterday_nav = franklin_mapping_sheet['G{}'.format(mapping_row)].value
                    try:
                        trade_shares = trade_value / yesterday_nav
                    except ZeroDivisionError:
                        trade_shares = 0
                    currentPeriodPnL = compiled_file_sheet['F{}'.format(row)].value
                    today_shares = franklin_mapping_sheet['H{}'.format(mapping_row)].value + trade_shares

                    portfolio_positions['A{}'.format(portfolio_positions_counter)].value = franklin_mapping_sheet['D{}'.format(mapping_row)].value + ' - ' + working_date_formatted # crystal reference
                    portfolio_positions['B{}'.format(portfolio_positions_counter)].value = franklin_mapping_sheet['D{}'.format(mapping_row)].value + ' Investment Pool'# investor acct
                    portfolio_positions['C{}'.format(portfolio_positions_counter)].value = franklin_mapping_sheet['C{}'.format(mapping_row)].value # ivstment acct
                    portfolio_positions['D{}'.format(portfolio_positions_counter)].value = today_shares # units closing
                    portfolio_positions['F{}'.format(portfolio_positions_counter)].value = compiled_file_sheet['G{}'.format(row)].value # investment closing bal
                    portfolio_positions['E{}'.format(portfolio_positions_counter)].value = portfolio_positions['F{}'.format(portfolio_positions_counter)].value / today_shares # closing nav
                    portfolio_positions['G{}'.format(portfolio_positions_counter)].value = 1 # invstor closing fx rate
                    portfolio_positions['H{}'.format(portfolio_positions_counter)].value = portfolio_positions['F{}'.format(portfolio_positions_counter)].value # invstor closing bal
                    portfolio_positions['I{}'.format(portfolio_positions_counter)].value = 1 # port closing fx
                    portfolio_positions['J{}'.format(portfolio_positions_counter)].value = portfolio_positions['F{}'.format(portfolio_positions_counter)].value # port closing bal

                    # fix addressed on line 729
                    # if 'ALT' in str(compiled_file_sheet['C{}'.format(row)].value):
                    #     alt_strat_fund += float(compiled_file_sheet['F{}'.format(row)].value)
                    # elif 'LS CRED' in str(compiled_file_sheet['C{}'.format(row)].value):
                    #     ls_fund += float(compiled_file_sheet['F{}'.format(row)].value)

                    # rewrite mapping file edit
                    franklin_mapping_sheet['F{}'.format(mapping_row)].value = working_date_datetime # update previous NAV
                    franklin_mapping_sheet['F{}'.format(mapping_row)].number_format = numbers.FORMAT_DATE_XLSX14
                    franklin_mapping_sheet['J{}'.format(mapping_row)].value = franklin_mapping_sheet['G{}'.format(mapping_row)].value # update previous NAV
                    franklin_mapping_sheet['K{}'.format(mapping_row)].value = franklin_mapping_sheet['H{}'.format(mapping_row)].value # update previous O/S
                    franklin_mapping_sheet['L{}'.format(mapping_row)].value = franklin_mapping_sheet['I{}'.format(mapping_row)].value # update previous Total Assets

                    franklin_mapping_sheet['G{}'.format(mapping_row)].value = portfolio_positions['E{}'.format(portfolio_positions_counter)].value # update for today NAV
                    franklin_mapping_sheet['H{}'.format(mapping_row)].value = today_shares # update for today O/S
                    franklin_mapping_sheet['I{}'.format(mapping_row)].value = portfolio_positions['F{}'.format(portfolio_positions_counter)].value # update for today Total Assets

                    portfolio_positions_counter += 1
                    franklin_update_counter += 1

    for mapping_row in range(2, franklin_mapping_sheet.max_row  + 1):
        for k2_row in range(2, k2_sheet.max_row + 1):
            if franklin_mapping_sheet['E{}'.format(mapping_row)].value == True and franklin_mapping_sheet['M{}'.format(mapping_row)].value == True and (str(k2_sheet['C{}'.format(k2_row)].value) + str(k2_sheet['D{}'.format(k2_row)].value)) == str(franklin_mapping_sheet['A{}'.format(mapping_row)].value):
                portfolio_accounts['A{}'.format(portfolio_accounts_counter)].value = franklin_mapping_sheet['D{}'.format(mapping_row)].value + ' - ' + working_date_formatted  # crystal ref
                portfolio_accounts['B{}'.format(portfolio_accounts_counter)].value = franklin_mapping_sheet['C{}'.format(mapping_row)].value  # Port Acct
                portfolio_accounts['C{}'.format(portfolio_accounts_counter)].value = None # opening Units
                portfolio_accounts['D{}'.format(portfolio_accounts_counter)].value = franklin_mapping_sheet['G{}'.format(mapping_row)].value # Opening Nav (Swung) Previous
                portfolio_accounts['E{}'.format(portfolio_accounts_counter)].value = None # Opening FX Rate
                portfolio_accounts['H{}'.format(portfolio_accounts_counter)].value = 0 # FX PnL
                portfolio_accounts['I{}'.format(portfolio_accounts_counter)].value = 0 # FX Hedge PnL
                portfolio_accounts['J{}'.format(portfolio_accounts_counter)].value = 0 # Interest income
                portfolio_accounts['K{}'.format(portfolio_accounts_counter)].value = 0 # int expense
                portfolio_accounts['L{}'.format(portfolio_accounts_counter)].value = 0 # admin fe
                portfolio_accounts['M{}'.format(portfolio_accounts_counter)].value = 0 # custody fee
                portfolio_accounts['N{}'.format(portfolio_accounts_counter)].value = 0 # misc expense
                portfolio_accounts['O{}'.format(portfolio_accounts_counter)].value = 0 # acct expenses
                portfolio_accounts['P{}'.format(portfolio_accounts_counter)].value = 0 # management fee
                portfolio_accounts['Q{}'.format(portfolio_accounts_counter)].value = 0 # perf fee
                portfolio_accounts['R{}'.format(portfolio_accounts_counter)].value = 0 # acct fees
                portfolio_accounts['S{}'.format(portfolio_accounts_counter)].value = k2_sheet['L{}'.format(k2_row)].value # closing bal
                portfolio_accounts['T{}'.format(portfolio_accounts_counter)].value = k2_sheet['N{}'.format(k2_row)].value # closing units
                portfolio_accounts['F{}'.format(portfolio_accounts_counter)].value = portfolio_accounts['D{}'.format(portfolio_accounts_counter)].value * portfolio_accounts['T{}'.format(portfolio_accounts_counter)].value # Adj Opening Bal
                portfolio_accounts['G{}'.format(portfolio_accounts_counter)].value = portfolio_accounts['S{}'.format(portfolio_accounts_counter)].value - portfolio_accounts['F{}'.format(portfolio_accounts_counter)].value # Invst PnL
                portfolio_accounts['U{}'.format(portfolio_accounts_counter)].value = k2_sheet['L{}'.format(k2_row)].value / k2_sheet['N{}'.format(k2_row)].value # closing nav
                portfolio_accounts['V{}'.format(portfolio_accounts_counter)].value = k2_sheet['L{}'.format(k2_row)].value / k2_sheet['M{}'.format(k2_row)].value # fx rate

                # rewrite mapping file edit
                franklin_mapping_sheet['F{}'.format(mapping_row)].value = working_date_datetime # update previous NAV
                franklin_mapping_sheet['F{}'.format(mapping_row)].number_format = numbers.FORMAT_DATE_XLSX14
                franklin_mapping_sheet['J{}'.format(mapping_row)].value = franklin_mapping_sheet['G{}'.format(mapping_row)].value # update previous NAV
                franklin_mapping_sheet['K{}'.format(mapping_row)].value = franklin_mapping_sheet['H{}'.format(mapping_row)].value # update previous O/S
                franklin_mapping_sheet['L{}'.format(mapping_row)].value = franklin_mapping_sheet['I{}'.format(mapping_row)].value # update previous Total Assets

                franklin_mapping_sheet['G{}'.format(mapping_row)].value = portfolio_accounts['U{}'.format(portfolio_accounts_counter)].value # update for today NAV
                franklin_mapping_sheet['H{}'.format(mapping_row)].value = portfolio_accounts['T{}'.format(portfolio_accounts_counter)].value # update for today O/S
                franklin_mapping_sheet['I{}'.format(mapping_row)].value = portfolio_accounts['S{}'.format(portfolio_accounts_counter)].value # update for today Total Assets

                portfolio_accounts_counter += 1
                franklin_update_counter += 1


    # initialize investment pools with their PnL
    investment_pool_funds = {}

    # create a dictionary correlating Investment Pool Name: Investment Pool PnL
    for row in range(2, portfolio_positions.max_row + 1):
        if portfolio_positions['B{}'.format(row)].value and portfolio_positions['B{}'.format(row)].value not in investment_pool_funds:
            investment_pool_funds.setdefault(str(portfolio_positions['B{}'.format(row)].value), 0)

    # comparing the id's for portfolio positions in compiled file sheet with mapping sheet
    # we will see if the Investor Name for XML positions is equal, thus adding PnL to the dictionary key
    for row in range(2, compiled_file_sheet.max_row + 1):
        for mapping_row in range(2, franklin_mapping_sheet.max_row):
            if compiled_file_sheet['B{}'.format(row)].value == franklin_mapping_sheet['A{}'.format(mapping_row)].value:
                # Column D in franklin mapping sheet is missing investment pool attached, thus using Str func to append
                investment_pool_funds[str(franklin_mapping_sheet['D{}'.format(mapping_row)].value) + ' Investment Pool'] += compiled_file_sheet['F{}'.format(row)].value

    # fix addressed above
    # investment_pool_funds = ['FTIF Franklin K2 Alternative Strategies Fund Investment Pool',
    #                         'FTIF Franklin K2 Long Short Credit Fund Investment Pool']
    #investment_pool_PnL = [alt_strat_fund, ls_fund]

    for num, key in enumerate(investment_pool_funds):
        current_row = franklin_mapping_sheet.max_row - len(investment_pool_funds) + 1 + num
        portfolio_accounts['A{}'.format(portfolio_accounts_counter)].value = str(franklin_mapping_sheet['D{}'.format(current_row)].value) + ' - ' + working_date_formatted # crystal ref
        portfolio_accounts['B{}'.format(portfolio_accounts_counter)].value = str(franklin_mapping_sheet['C{}'.format(current_row)].value) # Port Acct
        portfolio_accounts['C{}'.format(portfolio_accounts_counter)].value = None # opening Units
        portfolio_accounts['D{}'.format(portfolio_accounts_counter)].value = float(franklin_mapping_sheet['G{}'.format(current_row)].value) # Opening Nav
        portfolio_accounts['E{}'.format(portfolio_accounts_counter)].value = None # Opening FX Rate
        portfolio_accounts['F{}'.format(portfolio_accounts_counter)].value = franklin_mapping_sheet['G{}'.format(current_row)].value * franklin_mapping_sheet['H{}'.format(current_row)].value # Adj Opening Bal
        portfolio_accounts['G{}'.format(portfolio_accounts_counter)].value = investment_pool_funds[str(franklin_mapping_sheet['C{}'.format(current_row)].value)] # Invst PnL
        portfolio_accounts['H{}'.format(portfolio_accounts_counter)].value = 0 # FX PnL
        portfolio_accounts['I{}'.format(portfolio_accounts_counter)].value = 0 # FX Hedge PnL
        portfolio_accounts['J{}'.format(portfolio_accounts_counter)].value = 0 # Interest income
        portfolio_accounts['K{}'.format(portfolio_accounts_counter)].value = 0 # int expense
        portfolio_accounts['L{}'.format(portfolio_accounts_counter)].value = 0 # admin fee
        portfolio_accounts['M{}'.format(portfolio_accounts_counter)].value = 0 # custody fee
        portfolio_accounts['N{}'.format(portfolio_accounts_counter)].value = 0 # misc expense
        portfolio_accounts['O{}'.format(portfolio_accounts_counter)].value = 0 # acct expenses
        portfolio_accounts['P{}'.format(portfolio_accounts_counter)].value = 0 # management fee
        portfolio_accounts['Q{}'.format(portfolio_accounts_counter)].value = 0 # perf fee
        portfolio_accounts['R{}'.format(portfolio_accounts_counter)].value = 0 # acct fees
        portfolio_accounts['S{}'.format(portfolio_accounts_counter)].value = portfolio_accounts['F{}'.format(portfolio_accounts_counter)].value + portfolio_accounts['G{}'.format(portfolio_accounts_counter)].value # closing bal
        portfolio_accounts['T{}'.format(portfolio_accounts_counter)].value = franklin_mapping_sheet['H{}'.format(current_row)].value # closing units from mapping file
        portfolio_accounts['U{}'.format(portfolio_accounts_counter)].value = portfolio_accounts['S{}'.format(portfolio_accounts_counter)].value / portfolio_accounts['T{}'.format(portfolio_accounts_counter)].value # closing nav
        portfolio_accounts['V{}'.format(portfolio_accounts_counter)].value = 1

        # rewrite for mapping file edit
        franklin_mapping_sheet['F{}'.format(current_row)].value = working_date_datetime # update previous NAV
        franklin_mapping_sheet['F{}'.format(current_row)].number_format = numbers.FORMAT_DATE_XLSX14
        franklin_mapping_sheet['J{}'.format(current_row)].value = franklin_mapping_sheet['G{}'.format(current_row)].value # update previous NAV
        franklin_mapping_sheet['K{}'.format(current_row)].value = franklin_mapping_sheet['H{}'.format(current_row)].value # update previous O/S
        franklin_mapping_sheet['L{}'.format(current_row)].value = franklin_mapping_sheet['I{}'.format(current_row)].value # update previous Total Assets

        franklin_mapping_sheet['G{}'.format(current_row)].value = portfolio_accounts['U{}'.format(portfolio_accounts_counter)].value # update for today NAV
        franklin_mapping_sheet['H{}'.format(current_row)].value = portfolio_accounts['T{}'.format(portfolio_accounts_counter)].value # update for today O/S
        franklin_mapping_sheet['I{}'.format(current_row)].value = portfolio_accounts['S{}'.format(portfolio_accounts_counter)].value # update for today Total Assets

        portfolio_accounts_counter += 1
        franklin_update_counter += 1

    if (os=='Windows'):
        franklin_crystal_dest = r'G:/Shared drives/Operations/K2/JPM - Daily/Import/' + working_date_formatted + ' Franklin Portfolio Crystallization ' + todayStr + '.xlsx'
    else:
        franklin_crystal_dest = f'{folder_path}/output/' + working_date_formatted + ' Franklin Portfolio Crystallization ' + todayStr + '.xlsx'
    crystal.save(franklin_crystal_dest)

    # valuation file start
    if (os=='Windows'):
        valuation_file = r'G:/Shared drives/Operations/K2/JPM - Daily/Processing/Valuation Template.xlsx'
    else:
        valuation_file = f'{folder_path}/templates/Valuation Template.xlsx'
    valuation = openpyxl.load_workbook(valuation_file)
    mgr_accts_perf = valuation['Mgr Accts Performance']

    mgr_accts_perf.delete_rows(2, 1000) # clean out file

    valuation_counter = 2
    for mapping_row in range(2, ucits_mapping_sheet.max_row):
        if ucits_mapping_sheet['E{}'.format(mapping_row)].value == False:
            for row in range(2, compiled_file_sheet.max_row + 1):
                if compiled_file_sheet['B{}'.format(row)].value == ucits_mapping_sheet['A{}'.format(mapping_row)].value and ucits_mapping_sheet['H{}'.format(mapping_row)].value > 0 and ucits_mapping_sheet['O{}'.format(mapping_row)].value == True:
                    mgr_accts_perf['A{}'.format(valuation_counter)].value = ucits_mapping_sheet['C{}'.format(mapping_row)].value
                    mgr_accts_perf['B{}'.format(valuation_counter)].value = working_date_datetime # date
                    mgr_accts_perf['B{}'.format(valuation_counter)].number_format = numbers.FORMAT_DATE_XLSX14
                    mgr_accts_perf['D{}'.format(valuation_counter)].value = ucits_mapping_sheet['G{}'.format(mapping_row)].value
                    mgr_accts_perf['E{}'.format(valuation_counter)].value = 'TRUE'  # final
                    valuation_counter += 1

    # remove blank name
    for row in range(2, valuation_counter):
        if not mgr_accts_perf['A{}'.format(row)].value:
            mgr_accts_perf.delete_rows(row, 1)

    if (os=='Windows'):
        ucits_valuation_dest = r'G:/Shared drives/Operations/K2/JPM - Daily/Import/' + working_date_formatted + ' Valuation Data - Account Performance K2 JPM UCITS TD ' + todayStr + '.xlsx'
    else:
        ucits_valuation_dest = f'{folder_path}/output/' + working_date_formatted + ' Valuation Data - Account Performance K2 JPM UCITS TD ' + todayStr + '.xlsx'
    valuation.save(ucits_valuation_dest)

    # franklin valuation_file
    mgr_accts_perf.delete_rows(2, 1000) # clean out file

    valuation_counter = 2
    for mapping_row in range(2, franklin_mapping_sheet.max_row):
        if franklin_mapping_sheet['E{}'.format(mapping_row)].value == False:
            for row in range(2, compiled_file_sheet.max_row + 1):
                if compiled_file_sheet['B{}'.format(row)].value == franklin_mapping_sheet['A{}'.format(mapping_row)].value and franklin_mapping_sheet['H{}'.format(mapping_row)].value > 0 and franklin_mapping_sheet['O{}'.format(mapping_row)].value ==True:
                    mgr_accts_perf['A{}'.format(valuation_counter)].value = franklin_mapping_sheet['C{}'.format(mapping_row)].value
                    mgr_accts_perf['B{}'.format(valuation_counter)].value = working_date_datetime # date
                    mgr_accts_perf['B{}'.format(valuation_counter)].number_format = numbers.FORMAT_DATE_XLSX14
                    mgr_accts_perf['D{}'.format(valuation_counter)].value = franklin_mapping_sheet['G{}'.format(mapping_row)].value
                    mgr_accts_perf['E{}'.format(valuation_counter)].value = 'TRUE'  # final
                    valuation_counter += 1

    # remove blank name
    for row in range(2, valuation_counter):
        if not mgr_accts_perf['A{}'.format(row)].value:
            mgr_accts_perf.delete_rows(row, 1)

    if (os=='Windows'):
        franklin_valuation_dest = r'G:/Shared drives/Operations/K2/JPM - Daily/Import/' + working_date_formatted + ' Valuation Data - Account Performance K2 JPM Franklin TD ' + todayStr + '.xlsx'
    else:
        franklin_valuation_dest = f'{folder_path}/output/' + working_date_formatted + ' Valuation Data - Account Performance K2 JPM Franklin TD ' + todayStr + '.xlsx'
    valuation.save(franklin_valuation_dest)

    #transaction
    ##### transactions

    if (os=='Windows'):
        transaction_file = r'G:/Shared drives/Operations/K2/JPM - Daily/Processing/Transactions Template.xlsx'
    else:
        transaction_file = f'{folder_path}/templates/Transactions Template.xlsx'

    transaction = openpyxl.load_workbook(transaction_file)

    transaction_sheet = transaction['Transactions']

    UID = ucits_mapping_notes['B1'].value

    transaction_date = 'JPM UCITS Trx ' + working_date_datetime.strftime('%Y%m %d') + ' 0' + str(UID)

    # equity

    transaction_sheet.delete_rows(2, 3000)

    any_transactions = 0
    transaction_counter = 2

    for mapping_row in range(2, ucits_mapping_sheet.max_row):
        for ftaf_row in range(2, ftaf_sheet.max_row + 1):
            if ucits_mapping_sheet['H{}'.format(mapping_row)].value - ucits_mapping_sheet['K{}'.format(mapping_row)].value != 0 and ucits_mapping_sheet['C{}'.format(mapping_row)].value == ftaf_sheet['E{}'.format(ftaf_row)].value and ucits_mapping_sheet['N{}'.format(mapping_row)].value == True:
                transaction_sheet['A{}'.format(transaction_counter)].value = transaction_date # trans UID
                UID += 1
                transaction_date = 'JPM UCITS Trx ' + working_date_datetime.strftime('%Y%m %d') + ' 0' + str(UID)
                transaction_sheet['B{}'.format(transaction_counter)].value = ucits_mapping_sheet['B{}'.format(mapping_row)].value # invstor acct long name
                transaction_sheet['C{}'.format(transaction_counter)].value = ucits_mapping_sheet['C{}'.format(mapping_row)].value # invstment acct long name
                transaction_sheet['D{}'.format(transaction_counter)].value = working_date_datetime # transaction date
                transaction_sheet['D{}'.format(transaction_counter)].number_format = numbers.FORMAT_DATE_XLSX14
                transaction_sheet['E{}'.format(transaction_counter)].value = working_date_datetime # custody date
                transaction_sheet['E{}'.format(transaction_counter)].number_format = numbers.FORMAT_DATE_XLSX14
                if ucits_mapping_sheet['H{}'.format(mapping_row)].value - ucits_mapping_sheet['K{}'.format(mapping_row)].value > 0:
                    transaction_sheet['F{}'.format(transaction_counter)].value = 'Buy' # transaction type
                    transaction_sheet['G{}'.format(transaction_counter)].value = abs(ucits_mapping_sheet['H{}'.format(mapping_row)].value - ucits_mapping_sheet['K{}'.format(mapping_row)].value) #units
                else:
                    transaction_sheet['F{}'.format(transaction_counter)].value = 'Sell' # transaction type
                    transaction_sheet['G{}'.format(transaction_counter)].value = abs(ucits_mapping_sheet['H{}'.format(mapping_row)].value - ucits_mapping_sheet['K{}'.format(mapping_row)].value) #units
                #Anatoliy can you check the following line of code I don;t htink you changed the mapping sheet column to J
                transaction_sheet['H{}'.format(transaction_counter)].value = ucits_mapping_sheet['G{}'.format(mapping_row)].value # NAV
                transaction_sheet['I{}'.format(transaction_counter)].value = transaction_sheet['G{}'.format(transaction_counter)].value * transaction_sheet['H{}'.format(transaction_counter)].value # local invstment acct
                transaction_sheet['J{}'.format(transaction_counter)].value = 1 # exchange rate
                transaction_sheet['K{}'.format(transaction_counter)].value = transaction_sheet['I{}'.format(transaction_counter)].value # investor amt
                transaction_sheet['L{}'.format(transaction_counter)].value = 0 # equalization
                transaction_sheet['M{}'.format(transaction_counter)].value = 'Settled' # settled
                transaction_counter += 1
                any_transactions += 1


    if any_transactions > 0:
        if (os=='Windows'):
            transaction_dest = r'G:/Shared drives/Operations/K2/JPM - Daily/Import/' + working_date_formatted + ' Equity - Transactions K2 JPM UCITS TD ' + todayStr + '.xlsx'
        else:
            transaction_dest = f'{folder_path}/output/' + working_date_formatted + ' Equity - Transactions K2 JPM UCITS TD ' + todayStr + '.xlsx'
        transaction.save(transaction_dest)
        print("UCITS Equity transaction file created.\n")
    else:
        print("There are no UCITS Equity transactions.\n")


    # investments
    transaction_sheet.delete_rows(2, 3000)

    any_transactions = 0
    transaction_counter = 2

    for mapping_row in range(2, ucits_mapping_sheet.max_row):
        if ucits_mapping_sheet['N{}'.format(mapping_row)].value == True and ucits_mapping_sheet['H{}'.format(mapping_row)].value - ucits_mapping_sheet['K{}'.format(mapping_row)].value and 'Sleeve' in ucits_mapping_sheet['C{}'.format(mapping_row)].value:
            transaction_sheet['A{}'.format(transaction_counter)].value = transaction_date # trans UID
            UID += 1
            transaction_date = 'JPM UCITS Trx ' + working_date_datetime.strftime('%Y%m %d') + ' 0' + str(UID)
            transaction_sheet['B{}'.format(transaction_counter)].value = ucits_mapping_sheet['B{}'.format(mapping_row)].value # invstor acct long name
            transaction_sheet['C{}'.format(transaction_counter)].value = ucits_mapping_sheet['C{}'.format(mapping_row)].value # invstment acct long name
            transaction_sheet['D{}'.format(transaction_counter)].value = working_date_datetime # transaction date
            transaction_sheet['D{}'.format(transaction_counter)].number_format = numbers.FORMAT_DATE_XLSX14
            transaction_sheet['E{}'.format(transaction_counter)].value = working_date_datetime # custody date
            transaction_sheet['E{}'.format(transaction_counter)].number_format = numbers.FORMAT_DATE_XLSX14
            if ucits_mapping_sheet['H{}'.format(mapping_row)].value - ucits_mapping_sheet['K{}'.format(mapping_row)].value > 0:
                transaction_sheet['F{}'.format(transaction_counter)].value = 'Buy' # transaction type
                transaction_sheet['G{}'.format(transaction_counter)].value = abs(ucits_mapping_sheet['H{}'.format(mapping_row)].value - ucits_mapping_sheet['K{}'.format(mapping_row)].value) #units
            else:
                transaction_sheet['F{}'.format(transaction_counter)].value = 'Sell' # transaction type
                transaction_sheet['G{}'.format(transaction_counter)].value = abs(ucits_mapping_sheet['H{}'.format(mapping_row)].value - ucits_mapping_sheet['K{}'.format(mapping_row)].value) #units
            transaction_sheet['H{}'.format(transaction_counter)].value = ucits_mapping_sheet['J{}'.format(mapping_row)].value # NAV
            transaction_sheet['I{}'.format(transaction_counter)].value = transaction_sheet['G{}'.format(transaction_counter)].value * transaction_sheet['H{}'.format(transaction_counter)].value # local invstment acct
            transaction_sheet['J{}'.format(transaction_counter)].value = 1 # exchange rate
            transaction_sheet['K{}'.format(transaction_counter)].value = transaction_sheet['I{}'.format(transaction_counter)].value # investor amt
            transaction_sheet['L{}'.format(transaction_counter)].value = 0 # equalization
            transaction_sheet['M{}'.format(transaction_counter)].value = 'Settled' # settled
            transaction_counter += 1
            any_transactions += 1

    if any_transactions > 0:
        if (os=='Windows'):
            transaction_dest = r'G:/Shared drives/Operations/K2/JPM - Daily/Import/' + working_date_formatted + ' Investments - Transactions K2 JPM UCITS TD ' + todayStr + '.xlsx'
        else:
            transaction_dest = f'{folder_path}/output/' + working_date_formatted + ' Investments - Transactions K2 JPM UCITS TD ' + todayStr + '.xlsx'
        transaction.save(transaction_dest)
        print("UCITS Investments transaction file created.\n")
    else:
        print("There are no UCITS Investments transactions.\n")

    #franklin transaction_sheet

    UID = ucits_mapping_notes['B1'].value

    transaction_date = 'JPM Franklin Trx ' + working_date_datetime.strftime('%Y%m %d') + ' 0' + str(UID)

    # equity

    transaction_sheet.delete_rows(2, 3000)

    any_transactions = 0
    transaction_counter = 2

    for mapping_row in range(2, franklin_mapping_sheet.max_row):
        for k2_row in range(2, k2_sheet.max_row+1):
            if franklin_mapping_sheet['H{}'.format(mapping_row)].value - franklin_mapping_sheet['K{}'.format(mapping_row)].value != 0 and franklin_mapping_sheet['A{}'.format(mapping_row)].value == (str(k2_sheet['C{}'.format(k2_row)].value) + str(k2_sheet['D{}'.format(k2_row)].value)):
                transaction_sheet['A{}'.format(transaction_counter)].value = transaction_date # trans UID
                UID += 1
                transaction_date = 'JPM Franklin Trx ' + working_date_datetime.strftime('%Y%m %d') + ' 0' + str(UID)
                transaction_sheet['B{}'.format(transaction_counter)].value = franklin_mapping_sheet['B{}'.format(mapping_row)].value # invstor acct long name
                transaction_sheet['C{}'.format(transaction_counter)].value = franklin_mapping_sheet['C{}'.format(mapping_row)].value # invstment acct long name
                transaction_sheet['D{}'.format(transaction_counter)].value = working_date_datetime # transaction date
                transaction_sheet['D{}'.format(transaction_counter)].number_format = numbers.FORMAT_DATE_XLSX14
                transaction_sheet['E{}'.format(transaction_counter)].value = working_date_datetime # custody date
                transaction_sheet['E{}'.format(transaction_counter)].number_format = numbers.FORMAT_DATE_XLSX14
                if franklin_mapping_sheet['H{}'.format(mapping_row)].value - franklin_mapping_sheet['K{}'.format(mapping_row)].value > 0:
                    transaction_sheet['F{}'.format(transaction_counter)].value = 'Buy' # transaction type
                    transaction_sheet['G{}'.format(transaction_counter)].value = abs(franklin_mapping_sheet['H{}'.format(mapping_row)].value - franklin_mapping_sheet['K{}'.format(mapping_row)].value) #units
                else:
                    transaction_sheet['F{}'.format(transaction_counter)].value = 'Sell' # transaction type
                    transaction_sheet['G{}'.format(transaction_counter)].value = abs(franklin_mapping_sheet['H{}'.format(mapping_row)].value - franklin_mapping_sheet['K{}'.format(mapping_row)].value) #units
                transaction_sheet['H{}'.format(transaction_counter)].value = franklin_mapping_sheet['G{}'.format(mapping_row)].value # NAV
                transaction_sheet['I{}'.format(transaction_counter)].value = transaction_sheet['G{}'.format(transaction_counter)].value * transaction_sheet['H{}'.format(transaction_counter)].value # local invstment acct
                transaction_sheet['J{}'.format(transaction_counter)].value = 1 # exchange rate
                transaction_sheet['K{}'.format(transaction_counter)].value = transaction_sheet['I{}'.format(transaction_counter)].value # investor amt
                transaction_sheet['L{}'.format(transaction_counter)].value = 0 # equalization
                transaction_sheet['M{}'.format(transaction_counter)].value = 'Settled' # settled
                transaction_counter += 1
                any_transactions += 1


    if any_transactions > 0:
        if (os=='Windows'):
            transaction_dest = r'G:/Shared drives/Operations/K2/JPM - Daily/Import/' + working_date_formatted + ' Equity - Transactions K2 JPM Franklin TD ' + todayStr + '.xlsx'
        else:
            transaction_dest = f'{folder_path}/output/' + working_date_formatted + ' Equity - Transactions K2 JPM Franklin TD ' + todayStr + '.xlsx'
        transaction.save(transaction_dest)
        print("Franklin Equity transaction file created.\n")
    else:
        print("There are no Franklin Equity transactions.\n")


    # investment
    transaction_sheet.delete_rows(2, 3000)

    any_transactions = 0
    transaction_counter = 2

    for mapping_row in range(2, franklin_mapping_sheet.max_row):
    	if  franklin_mapping_sheet['N{}'.format(mapping_row)].value == True and franklin_mapping_sheet['H{}'.format(mapping_row)].value - franklin_mapping_sheet['K{}'.format(mapping_row)].value and 'Sleeve' in franklin_mapping_sheet['C{}'.format(mapping_row)].value:
    		transaction_sheet['A{}'.format(transaction_counter)].value = transaction_date # trans UID
    		UID += 1
    		transaction_date = 'JPM Franklin Trx ' + working_date_datetime.strftime('%Y%m %d') + ' 0' + str(UID)
    		transaction_sheet['B{}'.format(transaction_counter)].value = franklin_mapping_sheet['B{}'.format(mapping_row)].value # invstor acct long name
    		transaction_sheet['C{}'.format(transaction_counter)].value = franklin_mapping_sheet['C{}'.format(mapping_row)].value # invstment acct long name
    		transaction_sheet['D{}'.format(transaction_counter)].value = working_date_datetime # transaction date
    		transaction_sheet['D{}'.format(transaction_counter)].number_format = numbers.FORMAT_DATE_XLSX14
    		transaction_sheet['E{}'.format(transaction_counter)].value = working_date_datetime # custody date
    		transaction_sheet['E{}'.format(transaction_counter)].number_format = numbers.FORMAT_DATE_XLSX14
    		if franklin_mapping_sheet['H{}'.format(mapping_row)].value - franklin_mapping_sheet['K{}'.format(mapping_row)].value > 0:
    			transaction_sheet['F{}'.format(transaction_counter)].value = 'Buy' # transaction type
    			transaction_sheet['G{}'.format(transaction_counter)].value = abs(franklin_mapping_sheet['H{}'.format(mapping_row)].value - franklin_mapping_sheet['K{}'.format(mapping_row)].value) #units
    		else:
    			transaction_sheet['F{}'.format(transaction_counter)].value = 'Sell' # transaction type
    			transaction_sheet['G{}'.format(transaction_counter)].value = abs(franklin_mapping_sheet['H{}'.format(mapping_row)].value - franklin_mapping_sheet['K{}'.format(mapping_row)].value) #units
    		transaction_sheet['H{}'.format(transaction_counter)].value = franklin_mapping_sheet['J{}'.format(mapping_row)].value # NAV
    		transaction_sheet['I{}'.format(transaction_counter)].value = transaction_sheet['G{}'.format(transaction_counter)].value * transaction_sheet['H{}'.format(transaction_counter)].value # local invstment acct
    		transaction_sheet['J{}'.format(transaction_counter)].value = 1 # exchange rate
    		transaction_sheet['K{}'.format(transaction_counter)].value = transaction_sheet['I{}'.format(transaction_counter)].value # investor amt
    		transaction_sheet['L{}'.format(transaction_counter)].value = 0 # equalization
    		transaction_sheet['M{}'.format(transaction_counter)].value = 'Settled' # settled
    		transaction_counter += 1
    		any_transactions += 1

    if any_transactions > 0:
        if (os=='Windows'):
            transaction_dest = r'G:/Shared drives/Operations/K2/JPM - Daily/Import/' + working_date_formatted + ' Investments - Transactions K2 JPM Franklin TD ' + todayStr + '.xlsx'
        else:
            transaction_dest = f'{folder_path}/output/' + working_date_formatted + ' Investments - Transactions K2 JPM Franklin TD ' + todayStr + '.xlsx'
        transaction.save(transaction_dest)
        print("Franklin Investments transaction file created.\n")
    else:
        print("There are no Franklin Investments transactions.\n")

    #----------------------------------------------------------------------------
    # mapping file edit
    # rewrite over date, closing NAV, closing Units
    ucits_mapping_notes['B1'].value = 1
    ucits_mapping_notes['B2'].value = working_date_datetime
    ucits_mapping_notes.number_format = numbers.FORMAT_DATE_XLSX14

    print('There have been ' + str(ucits_update_counter) + ' updates to the UCITS mapping file.')

    ucits_mapping_file.save(ucits_mapping_file_ext)

    franklin_mapping_notes['B1'].value = 1
    franklin_mapping_notes['B2'].value = working_date_datetime
    franklin_mapping_notes.number_format = numbers.FORMAT_DATE_XLSX14

    print('There have been ' + str(franklin_update_counter) + ' updates to the Franklin mapping file.')

    franklin_mapping_file.save(franklin_mapping_file_ext)
